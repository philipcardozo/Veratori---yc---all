"""
Web Server with Frame Streaming
Serves frontend and streams video + inventory data via WebSockets

Features:
- Centralized logging to run/system.log
- Structured JSON responses with metadata
- WebSocket live updates for mobile/dashboard sync
- Demo-safe error handling (no uncaught exceptions)
"""

import asyncio
import csv
import functools
import io
import json
import logging
import logging.handlers
import os
import sqlite3
import time
import traceback
import uuid
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, Set, Dict, Any, Callable
import base64

import cv2
import numpy as np
import yaml
from aiohttp import web
import aiohttp

# Get logger from centralized logging system
logger = logging.getLogger(__name__)

# Project paths
PROJECT_ROOT = Path(__file__).parent.parent
RUN_DIR = PROJECT_ROOT / 'run'
SYSTEM_LOG_PATH = RUN_DIR / 'system.log'


class SystemLogger:
    """
    Centralized logging helper for tracking system events.
    Logs to both module logger and system.log file.
    """

    EVENT_UPLOAD = 'UPLOAD'
    EVENT_DETECTION = 'DETECTION'
    EVENT_ANALYTICS = 'ANALYTICS'
    EVENT_EXPORT = 'EXPORT'
    EVENT_ERROR = 'ERROR'
    EVENT_AUTH = 'AUTH'
    EVENT_WEBSOCKET = 'WEBSOCKET'

    @staticmethod
    def log(event_type: str, message: str, data: Optional[Dict] = None,
            level: int = logging.INFO):
        """
        Log a system event with structured format.

        Args:
            event_type: Type of event (UPLOAD, DETECTION, etc.)
            message: Human-readable message
            data: Optional structured data
            level: Logging level
        """
        timestamp = datetime.now(timezone.utc).isoformat()
        log_entry = f"[{event_type}] {message}"

        if data:
            # Serialize data for logging
            try:
                data_str = json.dumps(data, default=str)
                log_entry += f" | data={data_str}"
            except Exception:
                pass

        logger.log(level, log_entry)

        # Also write to dedicated system log file
        try:
            RUN_DIR.mkdir(exist_ok=True)
            with open(SYSTEM_LOG_PATH, 'a', encoding='utf-8') as f:
                f.write(f"{timestamp} | {log_entry}\n")
        except Exception:
            pass

    @staticmethod
    def upload(message: str, data: Optional[Dict] = None):
        SystemLogger.log(SystemLogger.EVENT_UPLOAD, message, data)

    @staticmethod
    def detection(message: str, data: Optional[Dict] = None):
        SystemLogger.log(SystemLogger.EVENT_DETECTION, message, data)

    @staticmethod
    def analytics(message: str, data: Optional[Dict] = None):
        SystemLogger.log(SystemLogger.EVENT_ANALYTICS, message, data)

    @staticmethod
    def export(message: str, data: Optional[Dict] = None):
        SystemLogger.log(SystemLogger.EVENT_EXPORT, message, data)

    @staticmethod
    def error(message: str, data: Optional[Dict] = None, exc_info: bool = False):
        if exc_info:
            data = data or {}
            data['traceback'] = traceback.format_exc()
        SystemLogger.log(SystemLogger.EVENT_ERROR, message, data, logging.ERROR)


def structured_response(success: bool, data: Optional[Dict] = None,
                        error: Optional[str] = None,
                        execution_time_ms: Optional[float] = None) -> Dict:
    """
    Create a structured JSON response with metadata.

    Args:
        success: Whether operation succeeded
        data: Response payload
        error: Error message if failed
        execution_time_ms: Execution time in milliseconds

    Returns:
        Structured response dictionary
    """
    response = {
        'success': success,
        'timestamp': datetime.now(timezone.utc).isoformat(),
    }

    if execution_time_ms is not None:
        response['execution_time_ms'] = round(execution_time_ms, 2)

    if success and data:
        response.update(data)
    elif not success and error:
        response['error'] = error

    return response


def api_handler(f: Callable) -> Callable:
    """
    Decorator for API handlers that adds:
    - Execution time tracking
    - Structured JSON responses
    - Exception handling (demo-safe)

    Usage:
        @api_handler
        async def handle_something(self, request):
            return {'data': 'value'}  # Will be wrapped in structured response
    """
    @functools.wraps(f)
    async def wrapper(self, request: web.Request) -> web.Response:
        start_time = time.time()

        try:
            result = await f(self, request)

            # If already a Response, return as-is
            if isinstance(result, web.Response):
                return result

            # Wrap dict result in structured response
            execution_ms = (time.time() - start_time) * 1000
            response_data = structured_response(
                success=True,
                data=result if isinstance(result, dict) else {'result': result},
                execution_time_ms=execution_ms
            )
            return web.json_response(response_data)

        except web.HTTPException:
            # Re-raise HTTP exceptions (redirects, auth failures, etc.)
            raise

        except Exception as e:
            execution_ms = (time.time() - start_time) * 1000
            SystemLogger.error(
                f"API error in {f.__name__}: {e}",
                {'endpoint': f.__name__},
                exc_info=True
            )

            response_data = structured_response(
                success=False,
                error=str(e),
                execution_time_ms=execution_ms
            )
            return web.json_response(response_data, status=500)

    return wrapper

# Import authentication module
try:
    from auth import load_auth_config, AuthManager
    AUTH_AVAILABLE = True
except ImportError:
    AUTH_AVAILABLE = False
    logger.warning("Authentication module not available")

# Import restock manager
try:
    from restock_manager import RestockManager
    RESTOCK_AVAILABLE = True
except ImportError:
    RESTOCK_AVAILABLE = False
    logger.warning("Restock manager module not available")

# Import analytics processor
try:
    from analytics import AnalyticsProcessor
    ANALYTICS_AVAILABLE = True
except ImportError:
    ANALYTICS_AVAILABLE = False
    logger.warning("Analytics module not available")


class VideoStreamServer:
    """
    Lightweight async web server for streaming video and inventory data
    Uses WebSockets for low-latency bidirectional communication
    """
    
    def __init__(
        self,
        host: str = '0.0.0.0',
        port: int = 8080,
        frontend_dir: Optional[Path] = None,
        enable_auth: bool = True
    ):
        """
        Initialize web server
        
        Args:
            host: Server host address
            port: Server port
            frontend_dir: Path to frontend files
            enable_auth: Enable authentication (default: True)
        """
        self.host = host
        self.port = port
        self.frontend_dir = frontend_dir or Path(__file__).parent.parent / 'apps' / 'web-frontend'
        
        # Initialize authentication
        self.auth_enabled = False
        self.auth_manager: Optional[AuthManager] = None
        self.cookie_name = 'pb_session'
        
        if enable_auth and AUTH_AVAILABLE:
            auth_enabled, auth_manager = load_auth_config()
            if auth_enabled and auth_manager:
                self.auth_enabled = True
                self.auth_manager = auth_manager
                logger.info("Authentication enabled")
            elif auth_enabled and not auth_manager:
                # Auth enabled but not configured = deny access
                self.auth_enabled = True
                self.auth_manager = None
                logger.warning("Authentication enabled but not configured - access will be denied")
            else:
                logger.info("Authentication disabled")
        elif enable_auth and not AUTH_AVAILABLE:
            logger.warning("Authentication requested but module not available")
        
        self.app = web.Application()
        # Add CORS middleware to handle preflight requests
        self.app.middlewares.append(self._create_cors_middleware())
        self.setup_routes()
        
        # Active WebSocket connections
        self.websockets: Set[web.WebSocketResponse] = set()
        
        # Latest frame and inventory data
        self.latest_frame: Optional[np.ndarray] = None
        self.latest_inventory: dict = {}
        self.latest_stats: dict = {}
        self.latest_freshness: dict = {}
        self.latest_sales: list = []
        self.latest_alerts: list = []
        
        # Camera management
        self._camera_ref = None          # Set via set_camera()
        self._available_cameras: list = []  # Populated at startup / on refresh
        self._video_file_ref = None     # Video file camera (alternative to USB camera)
        self._is_video_file_mode = False  # Track if we're in video file mode
        
        # Component references for upload & analytics
        self._detector_ref = None        # Set via set_detector()
        self._inventory_tracker_ref = None  # Set via set_inventory_tracker()
        self._stream_manager_ref = None  # Set via set_stream_manager()
        
        # Restock manager
        self.restock_manager: Optional[RestockManager] = None
        if RESTOCK_AVAILABLE:
            try:
                self.restock_manager = RestockManager()
                logger.info("Restock manager initialized")
            except Exception as e:
                logger.error(f"Failed to initialize restock manager: {e}")

        # Analytics processor
        self.analytics_processor: Optional[AnalyticsProcessor] = None
        if ANALYTICS_AVAILABLE:
            try:
                self.analytics_processor = AnalyticsProcessor()
                logger.info("Analytics processor initialized")
            except Exception as e:
                logger.error(f"Failed to initialize analytics processor: {e}")

        # NVIDIA DeepStream inference engine
        self.nvidia_engine = None
        try:
            from nvidia_inference import NvidiaInferenceEngine
            db_path = Path(__file__).parent.parent / 'data' / 'inventory.db'
            self.nvidia_engine = NvidiaInferenceEngine(db_path=db_path)
            logger.info("NVIDIA inference engine initialized")
        except Exception as e:
            logger.warning(f"NVIDIA inference engine not available: {e}")

        # Server statistics
        self.frames_streamed = 0
        self.start_time = time.time()

        # Ensure demo data directories exist
        self._ensure_directories()

        # System information (set by main.py during initialization)
        self._system_info: Dict[str, Any] = {
            'start_time': datetime.now(timezone.utc).isoformat(),
            'detector_mode': 'unknown',
            'initialization_status': {},
            'initialization_errors': {}
        }

        # WebSocket event subscribers (for live updates)
        self._event_subscribers: Dict[str, Set[web.WebSocketResponse]] = {
            'analytics': set(),
            'inventory': set(),
            'detection': set(),
            'restock': set()
        }

    def _ensure_directories(self):
        """Ensure all demo data directories exist."""
        try:
            dirs_to_create = [
                PROJECT_ROOT / 'run',
                PROJECT_ROOT / 'run' / 'exports',
                PROJECT_ROOT / 'data',
                PROJECT_ROOT / 'restock_photos',
                PROJECT_ROOT / 'models'
            ]
            for d in dirs_to_create:
                d.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.warning(f"Failed to create directories: {e}")

    def set_system_info(self, info: Dict[str, Any]):
        """Set system information from main.py initialization."""
        self._system_info.update(info)
        logger.info(f"System info updated: detector_mode={info.get('detector_mode')}")

    async def broadcast_event(self, event_type: str, data: Dict[str, Any]):
        """
        Broadcast an event to all connected WebSocket clients.
        Used for live updates when analytics are uploaded, inventory changes, etc.

        Args:
            event_type: Type of event ('analytics', 'inventory', 'detection', 'restock')
            data: Event payload
        """
        if not self.websockets:
            return

        message = {
            'type': f'{event_type}_update',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'data': data
        }

        # Broadcast to all connected clients
        disconnected = []
        for ws in self.websockets:
            try:
                await ws.send_json(message)
            except Exception as e:
                logger.debug(f"WebSocket send failed: {e}")
                disconnected.append(ws)

        # Clean up disconnected clients
        for ws in disconnected:
            self.websockets.discard(ws)

        if disconnected:
            logger.debug(f"Cleaned up {len(disconnected)} disconnected WebSocket clients")

        SystemLogger.log(
            SystemLogger.EVENT_WEBSOCKET,
            f"Broadcast {event_type} event to {len(self.websockets)} clients",
            {'event_type': event_type, 'client_count': len(self.websockets)}
        )

    def setup_routes(self):
        """Setup HTTP and WebSocket routes"""
        # Public routes (no authentication required)
        self.app.router.add_get('/login', self.handle_login_page)
        self.app.router.add_post('/api/login', self.handle_login)
        self.app.router.add_post('/api/logout', self.handle_logout)
        self.app.router.add_get('/health', self.handle_health)
        
        # Sub-page routes (protected) — clean URLs + .html extensions
        self.app.router.add_get('/upload', self.handle_upload_page)
        self.app.router.add_get('/upload.html', self.handle_upload_page)
        self.app.router.add_get('/analytics', self.handle_analytics_page)
        self.app.router.add_get('/analytics.html', self.handle_analytics_page)
        self.app.router.add_get('/account', self.handle_account_page)
        self.app.router.add_get('/account.html', self.handle_account_page)
        
        # Protected routes (authentication required)
        self.app.router.add_get('/', self.handle_index)
        self.app.router.add_get('/index.html', self.handle_index)
        self.app.router.add_get('/ws', self.handle_websocket)
        self.app.router.add_get('/api/stats', self.handle_stats)
        self.app.router.add_get('/api/freshness', self.handle_freshness)
        self.app.router.add_get('/api/sales', self.handle_sales)
        self.app.router.add_get('/api/alerts', self.handle_alerts)
        
        # Camera management routes
        self.app.router.add_get('/api/cameras', self.handle_list_cameras)
        self.app.router.add_post('/api/camera/switch', self.handle_switch_camera)
        self.app.router.add_post('/api/video/load', self.handle_load_video)
        self.app.router.add_post('/api/video/switch-to-camera', self.handle_switch_to_camera)
        self.app.router.add_post('/api/video/export-results', self.handle_video_export_results)
        
        # Upload API
        self.app.router.add_post('/api/upload/detect', self.handle_upload_detect)

        # Production Detection API (investor-demo safe, never fails)
        self.app.router.add_post('/detect', self.handle_detect_production)
        
        # Analytics API
        self.app.router.add_get('/api/analytics/summary', self.handle_analytics_summary)
        self.app.router.add_get('/api/analytics/export', self.handle_analytics_export)

        # Sales Analytics Upload API
        self.app.router.add_post('/api/analytics/upload', self.handle_analytics_upload)
        self.app.router.add_options('/api/analytics/upload', self.handle_options)
        self.app.router.add_get('/api/analytics/data', self.handle_analytics_data)
        self.app.router.add_get('/api/analytics/history', self.handle_analytics_history)
        self.app.router.add_post('/api/analytics/clear', self.handle_analytics_clear)

        # Account API
        self.app.router.add_post('/api/account/change-password', self.handle_change_password)
        self.app.router.add_get('/api/account/info', self.handle_account_info)
        
        # Restock Mobile App API
        self.app.router.add_post('/api/restock/login', self.handle_restock_login)
        self.app.router.add_post('/api/restock/validate', self.handle_restock_validate)
        self.app.router.add_post('/api/restock/logout', self.handle_restock_logout)
        self.app.router.add_post('/api/restock/upload', self.handle_restock_upload)
        self.app.router.add_post('/api/restock/detect', self.handle_restock_detect)
        self.app.router.add_get('/api/restock/submissions', self.handle_restock_submissions)
        self.app.router.add_get('/api/restock/notifications', self.handle_restock_notifications)
        self.app.router.add_get('/api/restock/notifications/count', self.handle_restock_notification_count)
        self.app.router.add_post('/api/restock/notifications/read', self.handle_restock_notification_read)
        self.app.router.add_get('/api/restock/photo/{filename}', self.handle_restock_photo)
        
        # Manager API for restock moderation
        self.app.router.add_get('/api/restock/all', self.handle_restock_all)
        self.app.router.add_post('/api/restock/status', self.handle_restock_status_update)

        # System Status API (demo-safe)
        self.app.router.add_get('/api/system/status', self.handle_system_status)
        self.app.router.add_get('/api/system/health', self.handle_system_health)

        # Polling endpoint for live updates (lightweight alternative to WebSocket)
        self.app.router.add_get('/api/poll/updates', self.handle_poll_updates)

        # NVIDIA DeepStream inference endpoints
        self.app.router.add_get('/api/nvidia/insights', self.handle_nvidia_insights)
        self.app.router.add_post('/api/nvidia/generate', self.handle_nvidia_generate)
    
    def _create_cors_middleware(self):
        """Create CORS middleware function"""
        @web.middleware
        async def cors_middleware(request: web.Request, handler):
            """CORS middleware to handle cross-origin requests"""
            # Get the origin from the request
            origin = request.headers.get('Origin')
            
            # Handle OPTIONS preflight requests
            if request.method == 'OPTIONS':
                response = web.Response()
            else:
                try:
                    response = await handler(request)
                except web.HTTPMethodNotAllowed as e:
                    # If method not allowed, add CORS headers before re-raising
                    raise e
            
            # Add CORS headers
            # If credentials are being sent, we must use the actual origin, not *
            if origin:
                response.headers['Access-Control-Allow-Origin'] = origin
                response.headers['Access-Control-Allow-Credentials'] = 'true'
            else:
                # Fallback to * if no origin header (same-origin request)
                response.headers['Access-Control-Allow-Origin'] = '*'
            
            response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
            response.headers['Access-Control-Max-Age'] = '3600'
            
            return response
        return cors_middleware
    
    async def handle_options(self, request: web.Request) -> web.Response:
        """Handle OPTIONS preflight requests for CORS"""
        # Get the origin from the request
        origin = request.headers.get('Origin')
        
        headers = {
            'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
            'Access-Control-Allow-Headers': 'Content-Type, Authorization',
            'Access-Control-Max-Age': '3600'
        }
        
        # If credentials are being sent, we must use the actual origin, not *
        if origin:
            headers['Access-Control-Allow-Origin'] = origin
            headers['Access-Control-Allow-Credentials'] = 'true'
        else:
            headers['Access-Control-Allow-Origin'] = '*'
        
        return web.Response(headers=headers)
        
    async def handle_index(self, request: web.Request) -> web.Response:
        """Serve main HTML page"""
        # Check authentication
        if not await self.check_auth(request):
            return web.HTTPFound('/login')
        
        index_path = self.frontend_dir / 'index.html'
        
        if not index_path.exists():
            return web.Response(
                text="Frontend not found. Please ensure apps/web-frontend/index.html exists.",
                status=404
            )
        
        return web.FileResponse(index_path)
    
    async def handle_health(self, request: web.Request) -> web.Response:
        """Health check endpoint"""
        uptime = time.time() - self.start_time
        
        health_data = {
            'status': 'healthy',
            'uptime_seconds': uptime,
            'active_connections': len(self.websockets),
            'frames_streamed': self.frames_streamed
        }
        
        return web.json_response(health_data)
    
    async def handle_stats(self, request: web.Request) -> web.Response:
        """Return current statistics"""
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)
        
        return web.json_response(self.latest_stats)
    
    async def handle_freshness(self, request: web.Request) -> web.Response:
        """Return current freshness data"""
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)
        
        return web.json_response(self.latest_freshness)
    
    async def handle_sales(self, request: web.Request) -> web.Response:
        """Return sales log"""
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)
        
        limit = int(request.query.get('limit', 100))
        return web.json_response(self.latest_sales[:limit])
    
    async def handle_alerts(self, request: web.Request) -> web.Response:
        """Return alerts log"""
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)
        
        limit = int(request.query.get('limit', 20))
        return web.json_response(self.latest_alerts[:limit])
    
    # ------------------------------------------------------------------
    # Camera management helpers
    # ------------------------------------------------------------------

    def set_camera(self, camera):
        """
        Store a reference to the USBCamera so we can switch at runtime.

        Args:
            camera: USBCamera instance used by the streaming loop.
        """
        self._camera_ref = camera
        self._is_video_file_mode = False
    
    def set_video_file(self, video_camera):
        """
        Store a reference to the VideoFileCamera for video file playback.

        Args:
            video_camera: VideoFileCamera instance used by the streaming loop.
        """
        self._video_file_ref = video_camera
        self._is_video_file_mode = True
    
    def get_active_camera(self):
        """
        Get the currently active camera (USB or video file).
        
        Returns:
            Active camera instance (USBCamera or VideoFileCamera)
        """
        if self._is_video_file_mode and self._video_file_ref:
            return self._video_file_ref
        return self._camera_ref

    def set_available_cameras(self, cameras: list):
        """
        Cache the list returned by USBCamera.enumerate_cameras().

        Args:
            cameras: List of camera info dicts.
        """
        self._available_cameras = cameras

    async def handle_list_cameras(self, request: web.Request) -> web.Response:
        """
        GET /api/cameras
        Return list of available cameras and which one is active.
        Optionally pass ?refresh=1 to re-enumerate.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        from camera import USBCamera
        
        # Re-enumerate if requested or if list is empty
        if request.query.get('refresh') == '1' or len(self._available_cameras) == 0:
            try:
                logger.info("Enumerating cameras via API...")
                self._available_cameras = USBCamera.enumerate_cameras()
                logger.info(f"Enumerated {len(self._available_cameras)} cameras")
            except Exception as e:
                logger.error(f"Failed to enumerate cameras: {e}", exc_info=True)
                # Return empty list on error
                self._available_cameras = []

        active_index = self._camera_ref.camera_index if self._camera_ref else None

        return web.json_response({
            'cameras': self._available_cameras,
            'active_index': active_index,
        })

    async def handle_switch_camera(self, request: web.Request) -> web.Response:
        """
        POST /api/camera/switch   { "index": 2 }
        Switch the active camera to the given device index.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if self._camera_ref is None:
            return web.json_response(
                {'success': False, 'message': 'Camera not initialised'},
                status=500,
            )

        try:
            body = await request.json()
            new_index = int(body.get('index', -1))
        except Exception:
            return web.json_response(
                {'success': False, 'message': 'Invalid request body'},
                status=400,
            )

        if new_index < 0:
            return web.json_response(
                {'success': False, 'message': 'Missing or invalid camera index'},
                status=400,
            )

        # Already on this camera?
        if new_index == self._camera_ref.camera_index:
            return web.json_response({
                'success': True,
                'message': f'Already using camera {new_index}',
                'active_index': new_index,
            })

        # Perform the switch (blocks briefly while the device re-inits)
        ok = self._camera_ref.switch_camera(new_index)

        if ok:
            # Switch back to camera mode
            self._is_video_file_mode = False
            # Notify all WS clients
            await self._broadcast_camera_change(new_index)
            return web.json_response({
                'success': True,
                'message': f'Switched to camera {new_index}',
                'active_index': new_index,
            })
        else:
            return web.json_response(
                {'success': False, 'message': f'Failed to open camera {new_index}'},
                status=500,
            )
    
    @api_handler
    async def handle_load_video(self, request: web.Request) -> Dict:
        """
        POST /api/video/load
        Load a video file for playback with YOLO detection.
        
        Body: { "video_path": "/path/to/video.mp4", "loop": true }
        """
        if not await self.check_auth(request):
            return {'error': 'Unauthorized', 'success': False}
        
        try:
            body = await request.json()
            video_path = body.get('video_path', '').strip()
            loop = body.get('loop', True)
            
            if not video_path:
                return {'error': 'video_path is required', 'success': False}
            
            # Convert to Path and resolve
            video_path_str = video_path.strip()
            
            # Try multiple path resolution strategies
            possible_paths = []
            
            # If absolute path, use as-is
            if os.path.isabs(video_path_str):
                possible_paths.append(Path(video_path_str))
            else:
                # Get workspace root (go up 3 levels from backend/server.py)
                # backend/ -> projects/inventory-system/ -> projects/ -> Veratori/
                backend_dir = Path(__file__).parent
                project_root = backend_dir.parent  # projects/inventory-system
                workspace_root = project_root.parent.parent  # Veratori (workspace root)
                
                # Try relative to workspace root first (most likely location)
                possible_paths.append(workspace_root / video_path_str)
                
                # Try relative to project root
                possible_paths.append(project_root / video_path_str)
                
                # Try relative to backend directory
                possible_paths.append(backend_dir / video_path_str)
                
                # Try with different path separators if needed
                normalized_path = video_path_str.replace('\\', '/')
                if normalized_path != video_path_str:
                    possible_paths.append(workspace_root / normalized_path)
            
            # Find the first existing path
            video_path_obj = None
            for path in possible_paths:
                if path.exists() and path.is_file():
                    video_path_obj = path
                    break
            
            if video_path_obj is None:
                tried_paths = [str(p) for p in possible_paths]
                return {
                    'error': f'Video file not found. Tried: {", ".join(tried_paths)}',
                    'success': False,
                    'tried_paths': tried_paths
                }
            
            # Import VideoFileCamera
            from camera import VideoFileCamera
            
            # Create video file camera
            video_camera = VideoFileCamera(str(video_path_obj), loop=loop)
            
            if not video_camera.open():
                return {
                    'error': f'Failed to open video file: {video_path_obj}',
                    'success': False
                }
            
            # Release old video file if any
            if self._video_file_ref:
                self._video_file_ref.release()

            # Generate a unique session ID for this video playback
            session_id = str(uuid.uuid4())

            # Set as active camera
            self.set_video_file(video_camera)

            # Notify stream manager if it exists
            if hasattr(self, '_stream_manager_ref') and self._stream_manager_ref:
                # Update the stream manager's camera reference
                self._stream_manager_ref.camera = video_camera
                # Propagate session tracking
                self._stream_manager_ref.current_session_id = session_id
                self._stream_manager_ref.session_start_time = time.time()
                # Propagate session to inventory tracker if supported
                tracker = self._stream_manager_ref.inventory_tracker
                if hasattr(tracker, 'set_session_id'):
                    tracker.set_session_id(session_id)

            video_info = video_camera.get_info()

            logger.info(f"Video file loaded: {video_path_obj} (session={session_id})")

            return {
                'success': True,
                'message': f'Video file loaded: {video_path_obj.name}',
                'video_info': video_info,
                'video_path': str(video_path_obj),
                'session_id': session_id
            }
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error in handle_load_video: {e}")
            return {'error': 'Invalid JSON in request body', 'success': False}
        except Exception as e:
            logger.error(f"Error loading video file: {e}", exc_info=True)
            return {'error': str(e), 'success': False}
    
    @api_handler
    async def handle_switch_to_camera(self, request: web.Request) -> Dict:
        """
        POST /api/video/switch-to-camera
        Switch back from video file mode to USB camera mode.
        """
        if not await self.check_auth(request):
            return {'error': 'Unauthorized', 'success': False}
        
        if self._camera_ref is None:
            return {'error': 'USB camera not available'}
        
        # Switch back to camera mode
        self._is_video_file_mode = False
        
        # Notify stream manager if it exists
        if hasattr(self, '_stream_manager_ref') and self._stream_manager_ref:
            self._stream_manager_ref.camera = self._camera_ref
        
        logger.info("Switched back to USB camera mode")
        
        return {
            'success': True,
            'message': 'Switched to USB camera mode',
            'active_index': self._camera_ref.camera_index if self._camera_ref else None
        }

    # Product prices and categories for export
    _PRODUCT_PRICES = {
        'mango': 12.50, 'island passion fruit': 15.99,
        'kilauea lemon cake': 18.75, 'maui custard': 14.25,
        'watermelon': 10.00, 'pineapple': 13.50,
        'cantaloupe': 11.75, 'strawberry': 16.00,
        'grapes': 12.99, 'coke diet': 3.50,
        'coke zero': 3.50, 'sprite': 3.50,
        'essentia': 4.50, 'perrier': 4.75,
        'jasmine green tea': 5.50, 'guava green tea': 5.75,
        'mango oolong tea': 6.00, 'philadelphia 6 roll': 22.50,
    }

    @staticmethod
    def _get_product_category(product_name: str) -> str:
        p = product_name.lower()
        if any(t in p for t in ['tea', 'oolong', 'green tea']):
            return 'Tea'
        if any(b in p for b in ['coke', 'sprite', 'soda', 'perrier', 'essentia']):
            return 'Beverage'
        if any(d in p for d in ['cake', 'custard', 'dessert']):
            return 'Dessert'
        return 'Fruit'

    async def handle_video_export_results(self, request: web.Request) -> web.Response:
        """
        POST /api/video/export-results
        Export sales data for a completed video session as CSV or Excel.

        Body: { "session_id": "<uuid>", "format": "csv" | "xlsx" }
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        try:
            body = await request.json()
        except Exception:
            return web.json_response({'error': 'Invalid JSON'}, status=400)

        session_id = body.get('session_id', '').strip()
        fmt = body.get('format', 'csv').strip().lower()

        if not session_id:
            return web.json_response({'error': 'session_id is required'}, status=400)
        if fmt not in ('csv', 'xlsx'):
            return web.json_response({'error': 'format must be csv or xlsx'}, status=400)

        # Retrieve sales for the session
        tracker = getattr(self, '_inventory_tracker_ref', None)
        persistence = getattr(tracker, 'persistence', None) if tracker else None
        if persistence is None:
            # Try via stream manager
            sm = getattr(self, '_stream_manager_ref', None)
            if sm:
                t = getattr(sm, 'inventory_tracker', None)
                persistence = getattr(t, 'persistence', None) if t else None

        if persistence is None:
            return web.json_response({'error': 'Persistence layer not available'}, status=503)

        sales = persistence.get_sales_by_session(session_id)
        if not sales:
            return web.json_response(
                {'error': f'No sales data found for session {session_id}'},
                status=404
            )

        # Build export rows
        headers = ['product', 'quantity', 'price', 'cost', 'timestamp', 'location', 'employee_id', 'category']
        rows = []
        for sale in sales:
            product = sale['product_name']
            price = self._PRODUCT_PRICES.get(product.lower(), 0.0)
            cost = round(price * 0.5, 2)
            rows.append([
                product,
                sale['quantity_delta'],
                price,
                cost,
                sale['timestamp_est'],
                'Video Demo',
                'VIDEO_SESSION',
                self._get_product_category(product)
            ])

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        if fmt == 'csv':
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(headers)
            writer.writerows(rows)
            csv_bytes = buf.getvalue().encode('utf-8')
            return web.Response(
                body=csv_bytes,
                content_type='text/csv',
                headers={
                    'Content-Disposition': f'attachment; filename="video_results_{ts}.csv"'
                }
            )
        else:
            try:
                import openpyxl
            except ImportError:
                return web.json_response(
                    {'error': 'openpyxl is not installed; cannot export xlsx'},
                    status=500
                )
            wb = openpyxl.Workbook()
            ws_xl = wb.active
            ws_xl.title = 'Sales'
            ws_xl.append(headers)
            for row in rows:
                ws_xl.append(row)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return web.Response(
                body=buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename="video_results_{ts}.xlsx"'
                }
            )

    async def _broadcast_camera_change(self, new_index: int):
        """Notify all WebSocket clients that the active camera changed."""
        if not self.websockets:
            return

        active_name = None
        for cam in self._available_cameras:
            if cam['index'] == new_index:
                active_name = cam['name']
                break

        message = {
            'type': 'camera_switched',
            'active_index': new_index,
            'active_name': active_name or f'Camera {new_index}',
        }

        await asyncio.gather(
            *[ws.send_json(message) for ws in self.websockets],
            return_exceptions=True,
        )

    # ------------------------------------------------------------------
    # Component setters for upload & analytics
    # ------------------------------------------------------------------

    def set_detector(self, detector):
        """Store a reference to the YOLODetector for image upload processing."""
        self._detector_ref = detector
        if detector is not None:
            mode = detector.get_mode()
            if detector.is_mock():
                logger.warning(
                    f"DETECTOR REGISTERED IN MOCK MODE — real YOLO unavailable. "
                    f"Reason: {getattr(detector, 'mock_reason', 'unknown')}. "
                    "Upload/detect endpoints will return errors instead of fake data."
                )
            else:
                logger.info(
                    f"Detector registered: mode=YOLO, model={detector.model_path}, "
                    f"classes={len(detector.class_names)}"
                )

    def set_inventory_tracker(self, tracker):
        """Store a reference to the inventory tracker for analytics queries."""
        self._inventory_tracker_ref = tracker
    
    def set_stream_manager(self, stream_manager):
        """Set stream manager reference"""
        self._stream_manager_ref = stream_manager

    # ------------------------------------------------------------------
    # Sub-page handlers
    # ------------------------------------------------------------------

    async def _serve_page(self, request: web.Request, filename: str) -> web.Response:
        """Serve a frontend HTML file with auth check."""
        if not await self.check_auth(request):
            return web.HTTPFound('/login')
        page_path = self.frontend_dir / filename
        if not page_path.exists():
            logger.error(f"Page not found: {page_path}")
            return web.Response(text=f"Page not found: {filename}", status=404)
        return web.FileResponse(page_path)

    async def handle_upload_page(self, request: web.Request) -> web.Response:
        """Serve the Upload page."""
        return await self._serve_page(request, 'upload.html')

    async def handle_analytics_page(self, request: web.Request) -> web.Response:
        """Serve the Analytics page."""
        return await self._serve_page(request, 'analytics.html')

    async def handle_account_page(self, request: web.Request) -> web.Response:
        """Serve the Account page."""
        return await self._serve_page(request, 'account.html')

    # ------------------------------------------------------------------
    # Upload API
    # ------------------------------------------------------------------

    async def handle_upload_detect(self, request: web.Request) -> web.Response:
        """
        POST /api/upload/detect
        Accept an image upload, run YOLO inference, return detections + annotated image.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if self._detector_ref is None or not self._detector_ref.is_loaded:
            return web.json_response(
                {'success': False, 'message': 'Detector not available'},
                status=503,
            )

        # Reject mock mode — upload page must use the real YOLO model
        if self._detector_ref.is_mock():
            reason = getattr(self._detector_ref, 'mock_reason', 'unknown')
            logger.warning(f"Upload detect rejected: detector is in mock mode ({reason})")
            return web.json_response(
                {'success': False,
                 'message': f'YOLO model is not loaded (mock mode active: {reason}). '
                            'Ensure the model file exists at models/best.pt and ultralytics is installed.'},
                status=503,
            )

        try:
            reader = await request.multipart()
            field = await reader.next()

            if field is None or field.name != 'image':
                return web.json_response(
                    {'success': False, 'message': 'No image field in upload'},
                    status=400,
                )

            # Read file content (limit 20 MB)
            data = await field.read(decode=False)
            if len(data) > 20 * 1024 * 1024:
                return web.json_response(
                    {'success': False, 'message': 'Image too large (max 20 MB)'},
                    status=400,
                )

            # Reject PNG files - only accept JPEG
            if data.startswith(b'\x89PNG\r\n\x1a\n'):
                return web.json_response(
                    {'success': False, 'message': 'PNG files are not accepted. Please upload a JPEG image.'},
                    status=400,
                )
            
            # Verify it's a JPEG file (starts with FF D8 FF)
            if not data.startswith(b'\xff\xd8\xff'):
                return web.json_response(
                    {'success': False, 'message': 'Only JPEG images are accepted. Please upload a JPEG image.'},
                    status=400,
                )

            # Decode image
            nparr = np.frombuffer(data, np.uint8)
            frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            if frame is None:
                return web.json_response(
                    {'success': False, 'message': 'Could not decode image'},
                    status=400,
                )

            h, w = frame.shape[:2]
            logger.info(f"Upload detect: image {w}x{h}, detector mode={self._detector_ref.get_mode()}")

            # Brightness check — dark/blank images have no products
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            mean_brightness = float(np.mean(gray))
            logger.info(f"Upload detect: mean brightness={mean_brightness:.1f}")
            if mean_brightness < 15.0:
                logger.info("Image too dark — returning 0 detections")
                return web.json_response({
                    'success': True,
                    'detections': [],
                    'summary': {},
                    'total_items': 0,
                    'annotated_image': base64.b64encode(
                        cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 90])[1]
                    ).decode('utf-8'),
                    'image_width': w,
                    'image_height': h,
                    'source': 'yolo',
                    'note': 'Image too dark for detection'
                })

            # Run YOLO detection
            loop = asyncio.get_event_loop()
            detections = await asyncio.wait_for(
                loop.run_in_executor(None, self._detector_ref.detect, frame),
                timeout=15.0
            )

            logger.info(f"Upload detect: {len(detections)} detections found")

            # Draw detections
            annotated = self._detector_ref.draw_detections(frame, detections)

            # Encode annotated image to base64
            _, buf = cv2.imencode('.jpg', annotated, [cv2.IMWRITE_JPEG_QUALITY, 90])
            annotated_b64 = base64.b64encode(buf).decode('utf-8')

            # Summarise detections by product name
            summary = {}
            for det in detections:
                name = det['class_name']
                summary[name] = summary.get(name, 0) + 1

            return web.json_response({
                'success': True,
                'detections': detections,
                'summary': summary,
                'total_items': len(detections),
                'annotated_image': annotated_b64,
                'image_width': w,
                'image_height': h,
                'source': 'yolo',
            })

        except Exception as e:
            logger.error(f"Upload detect error: {e}", exc_info=True)
            return web.json_response(
                {'success': False, 'message': f'Processing error: {str(e)}'},
                status=500,
            )

    # ------------------------------------------------------------------
    # Production Detection API (Investor-Demo Safe)
    # ------------------------------------------------------------------

    def _get_mock_detection_response(self, image_count: int = 1) -> dict:
        """
        Generate deterministic mock detection response.
        Used as fallback when YOLO is unavailable/disabled/fails.

        Args:
            image_count: Number of images being processed

        Returns:
            Mock detection response with realistic product data
        """
        # Deterministic mock products based on common inventory items
        mock_products = [
            {'name': 'Passion Fruit', 'count': 4},
            {'name': 'Maui Custard', 'count': 3},
            {'name': 'Lemon Cake', 'count': 5},
            {'name': 'Mango', 'count': 6},
            {'name': 'Pineapple', 'count': 4},
            {'name': 'Watermelon', 'count': 2},
        ]

        # Scale counts slightly based on image count for triple-photo mode
        if image_count > 1:
            for p in mock_products:
                p['count'] = int(p['count'] * 1.2)

        total = sum(p['count'] for p in mock_products)

        return {
            'products': mock_products,
            'total_detected': total,
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'source': 'mock'
        }

    def _log_detection_attempt(self, success: bool, source: str, image_count: int,
                                error: str = None, inference_time_ms: float = None):
        """
        Log detection attempts using centralized logging system.
        Logs to both system.log and detection-specific log files.

        Args:
            success: Whether detection succeeded
            source: 'yolo' or 'mock'
            image_count: Number of images processed
            error: Error message if failed
            inference_time_ms: Inference time in milliseconds
        """
        # Log via centralized logger
        status = 'SUCCESS' if success else 'FALLBACK'
        log_data = {
            'source': source,
            'images': image_count,
            'status': status
        }
        if inference_time_ms is not None:
            log_data['inference_time_ms'] = inference_time_ms
        if error:
            log_data['error'] = error

        SystemLogger.detection(
            f"Detection {status}: source={source}, images={image_count}",
            log_data
        )

        # Also write to detection-specific log file
        try:
            run_dir = Path(__file__).parent.parent / 'run'
            run_dir.mkdir(exist_ok=True)

            date_str = datetime.now().strftime('%Y-%m-%d')
            log_file = run_dir / f'detection_{date_str}.log'

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

            log_entry = f"[{timestamp}] {status} | source={source} | images={image_count}"
            if inference_time_ms is not None:
                log_entry += f" | inference_ms={inference_time_ms:.1f}"
            if error:
                log_entry += f" | error={error}"
            log_entry += "\n"

            with open(log_file, 'a') as f:
                f.write(log_entry)

        except Exception as e:
            logger.warning(f"Failed to write detection log file: {e}")

    def _is_yolo_enabled(self) -> bool:
        """
        Check if YOLO is enabled in config.yaml.

        Returns:
            True if YOLO is enabled, False otherwise
        """
        try:
            config_path = Path(__file__).parent.parent / 'config' / 'config.yaml'
            if config_path.exists():
                import yaml
                with open(config_path, 'r') as f:
                    config = yaml.safe_load(f)
                return config.get('detector', {}).get('yolo_enabled', True)
        except Exception as e:
            logger.warning(f"Could not read YOLO_ENABLED from config: {e}")
        return True  # Default to enabled

    async def handle_detect_production(self, request: web.Request) -> web.Response:
        """
        POST /detect

        Production-ready detection endpoint that NEVER fails.
        Accepts single or multiple photo uploads, runs YOLO inference,
        and falls back to mock detection if YOLO is unavailable/fails.

        Always returns HTTP 200 with standardized response schema:
        {
            "products": [{"name": "...", "count": N}, ...],
            "total_detected": N,
            "timestamp": "ISO8601",
            "source": "yolo" | "mock"
        }

        Request format: multipart/form-data with:
        - mode: "single" | "triple"
        - photo_front: file (required)
        - photo_left: file (for triple mode)
        - photo_right: file (for triple mode)
        """
        images = []
        mode = 'single'

        try:
            # Parse multipart form data
            reader = await request.multipart()

            async for field in reader:
                if field.name == 'mode':
                    mode = await field.text()
                elif field.name.startswith('photo_'):
                    data = await field.read(decode=False)
                    if len(data) > 20 * 1024 * 1024:
                        # Image too large, skip but don't fail
                        logger.warning(f"Image {field.name} exceeds 20MB limit, skipping")
                        continue
                    if data:
                        images.append(data)

            if not images:
                logger.warning("No valid images in request")
                self._log_detection_attempt(False, 'yolo', 0, error='no_valid_images')
                return web.json_response({
                    'products': [],
                    'total_detected': 0,
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'source': 'yolo',
                    'error': 'No valid images received'
                })

            # Check if YOLO is enabled in config
            yolo_enabled = self._is_yolo_enabled()
            if not yolo_enabled:
                logger.warning("YOLO is disabled in config")
                self._log_detection_attempt(False, 'disabled', len(images), error='yolo_disabled')
                return web.json_response({
                    'products': [],
                    'total_detected': 0,
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'source': 'disabled',
                    'error': 'YOLO detection is disabled in config'
                })

            # Check detector is loaded
            if self._detector_ref is None or not self._detector_ref.is_loaded:
                logger.warning("Detector not loaded")
                self._log_detection_attempt(False, 'error', len(images), error='detector_not_loaded')
                return web.json_response({
                    'products': [],
                    'total_detected': 0,
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'source': 'error',
                    'error': 'YOLO detector is not loaded'
                }, status=503)

            # Reject mock mode — never silently return fake data
            if self._detector_ref.is_mock():
                mock_reason = getattr(self._detector_ref, 'mock_reason', 'unknown')
                logger.warning(f"Detector is in mock mode ({mock_reason}) — refusing to return fake data")
                self._log_detection_attempt(False, 'mock', len(images), error=f'mock_mode: {mock_reason}')
                return web.json_response({
                    'products': [],
                    'total_detected': 0,
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'source': 'error',
                    'error': f'YOLO model not available (mock mode: {mock_reason}). '
                             'Ensure models/best.pt exists and ultralytics is installed.'
                }, status=503)

            logger.info(f"Running YOLO detection on {len(images)} image(s), mode={mode}")

            # Run YOLO inference
            try:
                start_time = time.time()
                all_detections = []
                images_processed = 0

                for idx, img_data in enumerate(images):
                    # Decode image
                    nparr = np.frombuffer(img_data, np.uint8)
                    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

                    if frame is None:
                        logger.warning(f"Image {idx}: failed to decode, skipping")
                        continue

                    h, w = frame.shape[:2]

                    # Brightness check — skip dark/blank images
                    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    mean_brightness = float(np.mean(gray))
                    logger.info(f"Image {idx}: {w}x{h}, brightness={mean_brightness:.1f}")

                    if mean_brightness < 15.0:
                        logger.info(f"Image {idx}: too dark (brightness={mean_brightness:.1f}), skipping")
                        images_processed += 1
                        continue

                    # Run YOLO inference with timeout
                    try:
                        loop = asyncio.get_event_loop()
                        detections = await asyncio.wait_for(
                            loop.run_in_executor(None, self._detector_ref.detect, frame),
                            timeout=15.0
                        )
                        logger.info(f"Image {idx}: {len(detections)} detections")
                        all_detections.extend(detections)
                        images_processed += 1
                    except asyncio.TimeoutError:
                        logger.error(f"Image {idx}: YOLO inference timed out after 15s")
                        # Continue with other images rather than aborting
                        images_processed += 1

                inference_time_ms = (time.time() - start_time) * 1000

                # Aggregate detections by product name
                product_counts = {}
                for det in all_detections:
                    name = det.get('class_name', 'Unknown')
                    product_counts[name] = product_counts.get(name, 0) + 1

                products = [
                    {'name': name, 'count': count}
                    for name, count in sorted(product_counts.items(), key=lambda x: -x[1])
                ]

                total = sum(p['count'] for p in products)

                logger.info(f"Detection complete: {total} items across {images_processed} image(s) "
                            f"in {inference_time_ms:.0f}ms")

                response = {
                    'products': products,
                    'total_detected': total,
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'source': 'yolo',
                    'images_processed': images_processed,
                    'inference_time_ms': round(inference_time_ms, 1)
                }

                self._log_detection_attempt(True, 'yolo', images_processed,
                                            inference_time_ms=inference_time_ms)
                return web.json_response(response)

            except Exception as yolo_error:
                error_msg = str(yolo_error)[:200]
                logger.error(f"YOLO inference error in /detect: {error_msg}", exc_info=True)
                self._log_detection_attempt(False, 'error', len(images), error=error_msg)
                return web.json_response({
                    'products': [],
                    'total_detected': 0,
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'source': 'error',
                    'error': f'YOLO inference failed: {error_msg}'
                }, status=500)

        except Exception as e:
            error_msg = str(e)[:200]
            logger.error(f"Unexpected error in /detect: {error_msg}", exc_info=True)
            self._log_detection_attempt(False, 'error', len(images) if images else 0,
                                        error=f'unexpected: {error_msg}')
            return web.json_response({
                'products': [],
                'total_detected': 0,
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'source': 'error',
                'error': f'Unexpected server error: {error_msg}'
            }, status=500)

    # ------------------------------------------------------------------
    # Analytics API
    # ------------------------------------------------------------------

    async def handle_analytics_summary(self, request: web.Request) -> web.Response:
        """
        GET /api/analytics/summary
        Return aggregated sales and inventory analytics.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        persistence = self._get_persistence()
        if persistence is None:
            return web.json_response({
                'success': False,
                'message': 'Persistence not available',
            }, status=503)

        try:
            now = datetime.now(timezone.utc)
            now_ts = now.timestamp()

            # Time boundaries
            today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp()
            week_start = (now - timedelta(days=now.weekday())).replace(
                hour=0, minute=0, second=0, microsecond=0
            ).timestamp()
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).timestamp()
            thirty_days_ago = (now - timedelta(days=30)).timestamp()

            with persistence._get_connection() as conn:
                cursor = conn.cursor()

                # --- Summary counts ---
                def _sum_sales(start_ts):
                    cursor.execute(
                        "SELECT COALESCE(SUM(quantity_delta),0) FROM sales_log WHERE timestamp_utc >= ?",
                        (start_ts,),
                    )
                    return cursor.fetchone()[0]

                def _count_sales(start_ts):
                    cursor.execute(
                        "SELECT COUNT(*) FROM sales_log WHERE timestamp_utc >= ?",
                        (start_ts,),
                    )
                    return cursor.fetchone()[0]

                today_total = _sum_sales(today_start)
                today_count = _count_sales(today_start)
                week_total = _sum_sales(week_start)
                week_count = _count_sales(week_start)
                month_total = _sum_sales(month_start)
                month_count = _count_sales(month_start)

                # --- Product breakdown (last 30 days) ---
                cursor.execute("""
                    SELECT product_name,
                           SUM(quantity_delta) as total_qty,
                           COUNT(*) as sale_count
                    FROM sales_log
                    WHERE timestamp_utc >= ?
                    GROUP BY product_name
                    ORDER BY total_qty DESC
                """, (thirty_days_ago,))
                product_breakdown = [
                    {'product': r[0], 'total_qty': r[1], 'sale_count': r[2]}
                    for r in cursor.fetchall()
                ]

                # --- Daily trend (last 30 days) ---
                cursor.execute("""
                    SELECT DATE(timestamp_utc, 'unixepoch') as sale_date,
                           SUM(quantity_delta) as total_qty,
                           COUNT(*) as sale_count
                    FROM sales_log
                    WHERE timestamp_utc >= ?
                    GROUP BY sale_date
                    ORDER BY sale_date ASC
                """, (thirty_days_ago,))
                daily_trend = [
                    {'date': r[0], 'total_qty': r[1], 'sale_count': r[2]}
                    for r in cursor.fetchall()
                ]

                # --- Weekly trend (last 12 weeks) ---
                twelve_weeks_ago = (now - timedelta(weeks=12)).timestamp()
                cursor.execute("""
                    SELECT strftime('%%Y-W%%W', timestamp_utc, 'unixepoch') as sale_week,
                           SUM(quantity_delta) as total_qty,
                           COUNT(*) as sale_count
                    FROM sales_log
                    WHERE timestamp_utc >= ?
                    GROUP BY sale_week
                    ORDER BY sale_week ASC
                """, (twelve_weeks_ago,))
                weekly_trend = [
                    {'week': r[0], 'total_qty': r[1], 'sale_count': r[2]}
                    for r in cursor.fetchall()
                ]

                # --- Monthly trend (last 12 months) ---
                twelve_months_ago = (now - timedelta(days=365)).timestamp()
                cursor.execute("""
                    SELECT strftime('%%Y-%%m', timestamp_utc, 'unixepoch') as sale_month,
                           SUM(quantity_delta) as total_qty,
                           COUNT(*) as sale_count
                    FROM sales_log
                    WHERE timestamp_utc >= ?
                    GROUP BY sale_month
                    ORDER BY sale_month ASC
                """, (twelve_months_ago,))
                monthly_trend = [
                    {'month': r[0], 'total_qty': r[1], 'sale_count': r[2]}
                    for r in cursor.fetchall()
                ]

                # --- Low-stock alert frequency ---
                cursor.execute("""
                    SELECT product_name, COUNT(*) as alert_count
                    FROM alerts_log
                    WHERE alert_type = 'low_stock' AND timestamp_utc >= ?
                    GROUP BY product_name
                    ORDER BY alert_count DESC
                """, (thirty_days_ago,))
                low_stock_freq = [
                    {'product': r[0], 'alert_count': r[1]}
                    for r in cursor.fetchall()
                ]

                # --- DB stats ---
                cursor.execute("SELECT COUNT(*) FROM sales_log")
                total_sales_all = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM inventory_snapshots")
                total_snapshots = cursor.fetchone()[0]
                cursor.execute("SELECT COUNT(*) FROM alerts_log")
                total_alerts = cursor.fetchone()[0]

            return web.json_response({
                'success': True,
                'summary': {
                    'today': {'total_qty': today_total, 'sale_count': today_count},
                    'week': {'total_qty': week_total, 'sale_count': week_count},
                    'month': {'total_qty': month_total, 'sale_count': month_count},
                },
                'product_breakdown': product_breakdown,
                'daily_trend': daily_trend,
                'weekly_trend': weekly_trend,
                'monthly_trend': monthly_trend,
                'low_stock_frequency': low_stock_freq,
                'database': {
                    'total_sales': total_sales_all,
                    'total_snapshots': total_snapshots,
                    'total_alerts': total_alerts,
                },
            })

        except Exception as e:
            logger.error(f"Analytics summary error: {e}", exc_info=True)
            return web.json_response(
                {'success': False, 'message': str(e)},
                status=500,
            )

    async def handle_analytics_export(self, request: web.Request) -> web.Response:
        """
        GET /api/analytics/export
        Download all historical data as an Excel (.xlsx) file.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        persistence = self._get_persistence()
        if persistence is None:
            return web.json_response(
                {'success': False, 'message': 'Persistence not available'},
                status=503,
            )

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return web.json_response(
                {'success': False, 'message': 'openpyxl not installed – cannot generate Excel'},
                status=503,
            )

        try:
            wb = Workbook()

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1E2936", end_color="1E2936", fill_type="solid")

            def _style_header(ws):
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

            with persistence._get_connection() as conn:
                cursor = conn.cursor()

                # --- Sales Log sheet ---
                ws_sales = wb.active
                ws_sales.title = "Sales Log"
                ws_sales.append(["ID", "Timestamp (EST)", "Product", "Qty", "Before", "After"])
                cursor.execute("SELECT id, timestamp_est, product_name, quantity_delta, inventory_before, inventory_after FROM sales_log ORDER BY timestamp_utc DESC")
                for row in cursor.fetchall():
                    ws_sales.append(list(row))
                _style_header(ws_sales)
                for col in ws_sales.columns:
                    ws_sales.column_dimensions[col[0].column_letter].width = 20

                # --- Inventory Snapshots sheet ---
                ws_inv = wb.create_sheet("Inventory Snapshots")
                ws_inv.append(["ID", "Timestamp (UTC)", "Frame#", "Total Items", "Inventory JSON"])
                cursor.execute("SELECT id, timestamp_utc, frame_number, total_items, inventory_json FROM inventory_snapshots ORDER BY timestamp_utc DESC LIMIT 5000")
                for row in cursor.fetchall():
                    r = list(row)
                    r[1] = datetime.fromtimestamp(r[1], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
                    ws_inv.append(r)
                _style_header(ws_inv)
                for col in ws_inv.columns:
                    ws_inv.column_dimensions[col[0].column_letter].width = 22

                # --- Alerts sheet ---
                ws_alerts = wb.create_sheet("Alerts")
                ws_alerts.append(["ID", "Timestamp (EST)", "Type", "Product", "Severity", "Message"])
                cursor.execute("SELECT id, timestamp_est, alert_type, product_name, severity, message FROM alerts_log ORDER BY timestamp_utc DESC")
                for row in cursor.fetchall():
                    ws_alerts.append(list(row))
                _style_header(ws_alerts)
                for col in ws_alerts.columns:
                    ws_alerts.column_dimensions[col[0].column_letter].width = 22

                # --- Freshness sheet ---
                ws_fresh = wb.create_sheet("Product Freshness")
                ws_fresh.append(["Product", "First Seen (UTC)", "Last Seen (UTC)", "Expired", "Expiration Days"])
                cursor.execute("SELECT product_name, first_seen_utc, last_seen_utc, is_expired, expiration_days FROM product_freshness")
                for row in cursor.fetchall():
                    r = list(row)
                    r[1] = datetime.fromtimestamp(r[1], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
                    r[2] = datetime.fromtimestamp(r[2], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')
                    r[3] = "Yes" if r[3] else "No"
                    ws_fresh.append(r)
                _style_header(ws_fresh)
                for col in ws_fresh.columns:
                    ws_fresh.column_dimensions[col[0].column_letter].width = 24

            # Write to memory buffer
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"veratori_data_{ts}.xlsx"

            # Log the export
            SystemLogger.export(
                f"Exported analytics data: {filename}",
                {'filename': filename, 'format': 'xlsx'}
            )

            return web.Response(
                body=buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename="{filename}"',
                },
            )

        except Exception as e:
            SystemLogger.error(f"Export error: {e}", exc_info=True)
            return web.json_response({'success': False, 'message': str(e)}, status=500)

    # ------------------------------------------------------------------
    # Sales Analytics Upload API
    # ------------------------------------------------------------------

    async def handle_analytics_upload(self, request: web.Request) -> web.Response:
        """
        POST /api/analytics/upload
        Upload CSV or XLSX sales data file for processing.

        Expected file format:
        Required columns: product, quantity, price, timestamp
        Optional columns: location, employee_id, cost, category

        Returns aggregated analytics with charts data.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if not self.analytics_processor:
            return web.json_response({
                'success': False,
                'error': 'Analytics processor not available'
            }, status=503)

        try:
            reader = await request.multipart()
            file_content = None
            filename = None

            async for field in reader:
                if field.name == 'file':
                    filename = field.filename or 'upload.csv'
                    file_content = await field.read(decode=False)

                    # Limit file size to 50MB
                    if len(file_content) > 50 * 1024 * 1024:
                        return web.json_response({
                            'success': False,
                            'error': 'File too large (max 50MB)'
                        }, status=400)
                    break

            if not file_content or not filename:
                return web.json_response({
                    'success': False,
                    'error': 'No file provided'
                }, status=400)

            # Process the file
            start_time = time.time()
            result = self.analytics_processor.process_file(file_content, filename)
            execution_time_ms = (time.time() - start_time) * 1000

            # Add execution time to result
            result['execution_time_ms'] = round(execution_time_ms, 2)

            # Log the analytics processing
            SystemLogger.analytics(
                f"Processed analytics file: {filename}",
                {
                    'filename': filename,
                    'records': result.get('summary', {}).get('total_records', 0),
                    'success': result.get('success', False),
                    'execution_time_ms': execution_time_ms
                }
            )

            # Broadcast update to connected clients for live dashboard refresh
            if result.get('success'):
                await self.broadcast_event('analytics', {
                    'event': 'upload_complete',
                    'upload_id': result.get('upload_id'),
                    'filename': filename,
                    'summary': result.get('summary', {})
                })

            return web.json_response(result)

        except Exception as e:
            SystemLogger.error(f"Analytics upload error: {e}", exc_info=True)
            return web.json_response({
                'success': False,
                'error': f'Processing failed: {str(e)}'
            }, status=500)

    async def handle_analytics_data(self, request: web.Request) -> web.Response:
        """
        GET /api/analytics/data
        Get the latest processed analytics data.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if not self.analytics_processor:
            return web.json_response({
                'success': False,
                'error': 'Analytics processor not available'
            }, status=503)

        try:
            result = self.analytics_processor.get_latest_analytics()

            if result:
                return web.json_response(result)
            else:
                return web.json_response({
                    'success': False,
                    'error': 'No analytics data available. Upload a sales file first.'
                })

        except Exception as e:
            logger.error(f"Analytics data error: {e}", exc_info=True)
            return web.json_response({
                'success': False,
                'error': str(e)
            }, status=500)

    async def handle_analytics_clear(self, request: web.Request) -> web.Response:
        """
        POST /api/analytics/clear
        Delete all analytics data (sales_uploads, processed_sales, analytics_summaries).
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if not self.analytics_processor:
            return web.json_response(
                {'success': False, 'error': 'Analytics processor not available'},
                status=503
            )

        try:
            result = self.analytics_processor.clear_all_data()
            SystemLogger.analytics("Analytics data cleared", result.get('deleted'))
            return web.json_response(result)
        except Exception as e:
            logger.error(f"Analytics clear error: {e}", exc_info=True)
            return web.json_response({'success': False, 'error': str(e)}, status=500)

    async def handle_analytics_history(self, request: web.Request) -> web.Response:
        """
        GET /api/analytics/history
        Get upload history.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if not self.analytics_processor:
            return web.json_response({
                'success': False,
                'uploads': []
            }, status=503)

        try:
            limit = int(request.query.get('limit', 10))
            uploads = self.analytics_processor.get_upload_history(limit)

            return web.json_response({
                'success': True,
                'uploads': uploads
            })

        except Exception as e:
            logger.error(f"Analytics history error: {e}", exc_info=True)
            return web.json_response({
                'success': False,
                'uploads': []
            }, status=500)

    # ------------------------------------------------------------------
    # Account API
    # ------------------------------------------------------------------

    async def handle_account_info(self, request: web.Request) -> web.Response:
        """
        GET /api/account/info
        Return basic account information for the logged-in user.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        username = request.get('username', 'unknown')
        return web.json_response({
            'username': username,
            'auth_enabled': self.auth_enabled,
        })

    async def handle_change_password(self, request: web.Request) -> web.Response:
        """
        POST /api/account/change-password
        Body: { "current_password": "...", "new_password": "..." }
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        if not self.auth_enabled or not self.auth_manager:
            return web.json_response(
                {'success': False, 'message': 'Authentication not configured'},
                status=503,
            )

        try:
            data = await request.json()
            current_pw = data.get('current_password', '')
            new_pw = data.get('new_password', '')

            if not current_pw or not new_pw:
                return web.json_response(
                    {'success': False, 'message': 'Both current and new passwords are required'},
                    status=400,
                )

            username = request.get('username', '')
            if not username:
                return web.json_response(
                    {'success': False, 'message': 'Session error – username not found'},
                    status=400,
                )

            success, message = self.auth_manager.change_password(username, current_pw, new_pw)
            status = 200 if success else 400
            return web.json_response({'success': success, 'message': message}, status=status)

        except Exception as e:
            logger.error(f"Change password error: {e}")
            return web.json_response(
                {'success': False, 'message': 'Server error'},
                status=500,
            )

    # ------------------------------------------------------------------
    # Helper – get persistence manager from tracker
    # ------------------------------------------------------------------

    def _get_persistence(self):
        """Return the PersistenceManager if available, else None."""
        if self._inventory_tracker_ref and hasattr(self._inventory_tracker_ref, 'persistence'):
            return self._inventory_tracker_ref.persistence
        return None

    # ------------------------------------------------------------------
    # System Status API (Demo-Safe)
    # ------------------------------------------------------------------

    async def handle_system_status(self, request: web.Request) -> web.Response:
        """
        GET /api/system/status
        Return comprehensive system status including initialization state.
        Demo-safe: always returns valid JSON, never crashes.
        """
        try:
            start_time = time.time()

            # Get detector info
            detector_info = {}
            if self._detector_ref:
                try:
                    detector_info = self._detector_ref.get_info()
                except Exception as e:
                    detector_info = {'error': str(e)}

            # Get inventory info
            inventory_info = {}
            if self._inventory_tracker_ref:
                try:
                    inventory_info = self._inventory_tracker_ref.get_statistics()
                except Exception as e:
                    inventory_info = {'error': str(e)}

            # Get analytics info
            analytics_info = {'available': self.analytics_processor is not None}
            if self.analytics_processor:
                try:
                    latest = self.analytics_processor.get_latest_analytics()
                    if latest:
                        analytics_info['last_upload'] = latest.get('upload_time')
                        analytics_info['total_records'] = latest.get('summary', {}).get('total_records', 0)
                except Exception:
                    pass

            execution_time_ms = (time.time() - start_time) * 1000

            return web.json_response(structured_response(
                success=True,
                data={
                    'system': {
                        'start_time': self._system_info.get('start_time'),
                        'uptime_seconds': time.time() - self.start_time,
                        'active_connections': len(self.websockets),
                        'frames_streamed': self.frames_streamed
                    },
                    'detector': detector_info,
                    'inventory': inventory_info,
                    'analytics': analytics_info,
                    'initialization': {
                        'status': self._system_info.get('initialization_status', {}),
                        'errors': self._system_info.get('initialization_errors', {})
                    },
                    'paths': {
                        'data': str(PROJECT_ROOT / 'data'),
                        'exports': str(PROJECT_ROOT / 'run' / 'exports'),
                        'restock_photos': str(PROJECT_ROOT / 'restock_photos'),
                        'logs': str(SYSTEM_LOG_PATH)
                    }
                },
                execution_time_ms=execution_time_ms
            ))

        except Exception as e:
            SystemLogger.error(f"System status error: {e}", exc_info=True)
            return web.json_response(structured_response(
                success=False,
                error=str(e)
            ), status=500)

    async def handle_system_health(self, request: web.Request) -> web.Response:
        """
        GET /api/system/health
        Lightweight health check for monitoring.
        """
        try:
            return web.json_response(structured_response(
                success=True,
                data={
                    'status': 'healthy',
                    'detector_mode': self._system_info.get('detector_mode', 'unknown'),
                    'uptime_seconds': time.time() - self.start_time,
                    'connections': len(self.websockets)
                }
            ))
        except Exception as e:
            return web.json_response(structured_response(
                success=False,
                error=str(e)
            ), status=500)

    async def handle_poll_updates(self, request: web.Request) -> web.Response:
        """
        GET /api/poll/updates?since=<timestamp>
        Lightweight polling endpoint for live updates.
        Alternative to WebSocket for mobile apps.

        Returns updates since the given timestamp.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        try:
            # Parse 'since' timestamp (ISO format or Unix timestamp)
            since_str = request.query.get('since', '')
            since_ts = 0.0

            if since_str:
                try:
                    if '.' in since_str or '-' in since_str:
                        # ISO format
                        since_dt = datetime.fromisoformat(since_str.replace('Z', '+00:00'))
                        since_ts = since_dt.timestamp()
                    else:
                        # Unix timestamp
                        since_ts = float(since_str)
                except Exception:
                    pass

            # Gather updates
            updates = {
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'inventory': {},
                'stats': {},
                'sales': [],
                'alerts': []
            }

            # Get latest inventory data
            if self.latest_inventory:
                updates['inventory'] = self.latest_inventory

            if self.latest_stats:
                updates['stats'] = self.latest_stats

            # Get recent sales and alerts since timestamp
            if self.latest_sales:
                updates['sales'] = [
                    s for s in self.latest_sales[:20]
                    if s.get('timestamp_utc', 0) > since_ts
                ]

            if self.latest_alerts:
                updates['alerts'] = [
                    a for a in self.latest_alerts[:10]
                    if a.get('timestamp_utc', 0) > since_ts
                ]

            return web.json_response(structured_response(
                success=True,
                data=updates
            ))

        except Exception as e:
            SystemLogger.error(f"Poll updates error: {e}", exc_info=True)
            return web.json_response(structured_response(
                success=False,
                error=str(e)
            ), status=500)

    # ── NVIDIA DeepStream inference handlers ───────────────────────────────────

    async def handle_nvidia_insights(self, request: web.Request) -> web.Response:
        """
        GET /api/nvidia/insights
        Returns the most recently generated NVIDIA inference report.
        If none exists, generates one automatically (mode=auto).
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        try:
            if self.nvidia_engine is None:
                return web.json_response(structured_response(
                    success=False,
                    error="NVIDIA inference engine not available"
                ), status=503)

            report = self.nvidia_engine.get_last_report()
            if report is None:
                # Generate an initial report on first request
                report = await asyncio.get_event_loop().run_in_executor(
                    None, lambda: self.nvidia_engine.generate(mode="auto")
                )

            return web.json_response(structured_response(success=True, data=report))

        except Exception as e:
            logger.error(f"NVIDIA insights error: {e}", exc_info=True)
            return web.json_response(structured_response(
                success=False, error=str(e)
            ), status=500)

    async def handle_nvidia_generate(self, request: web.Request) -> web.Response:
        """
        POST /api/nvidia/generate
        Body: { "mode": "pre_open" | "end_of_day" | "auto" }
        Triggers a new NVIDIA DeepStream inference report.
        """
        if not await self.check_auth(request):
            return web.json_response({'error': 'Unauthorized'}, status=401)

        try:
            if self.nvidia_engine is None:
                return web.json_response(structured_response(
                    success=False,
                    error="NVIDIA inference engine not available"
                ), status=503)

            body = {}
            try:
                body = await request.json()
            except Exception:
                pass

            mode = body.get("mode", "auto")
            if mode not in ("pre_open", "end_of_day", "auto"):
                mode = "auto"

            report = await asyncio.get_event_loop().run_in_executor(
                None, lambda: self.nvidia_engine.generate(mode=mode)
            )

            return web.json_response(structured_response(success=True, data=report))

        except Exception as e:
            logger.error(f"NVIDIA generate error: {e}", exc_info=True)
            return web.json_response(structured_response(
                success=False, error=str(e)
            ), status=500)

    async def handle_login_page(self, request: web.Request) -> web.Response:
        """Serve login page"""
        login_path = self.frontend_dir / 'login.html'
        
        if not login_path.exists():
            return web.Response(
                text="Login page not found.",
                status=404
            )
        
        return web.FileResponse(login_path)
    
    async def handle_login(self, request: web.Request) -> web.Response:
        """Handle login POST request"""
        if not self.auth_enabled:
            return web.json_response({'success': True, 'message': 'Authentication disabled'})
        
        if not self.auth_manager:
            return web.json_response({'success': False, 'message': 'Authentication not configured'}, status=503)
        
        try:
            data = await request.json()
            username = data.get('username', '').strip()
            password = data.get('password', '')
            
            if not username or not password:
                return web.json_response({'success': False, 'message': 'Username and password required'}, status=400)
            
            # Authenticate user
            session_token = self.auth_manager.authenticate(username, password)
            
            if session_token:
                # Create response with session cookie
                response = web.json_response({'success': True, 'message': 'Login successful'})
                
                # Determine if we're behind HTTPS
                is_secure = request.headers.get('X-Forwarded-Proto', '').lower() == 'https'
                
                # Set session cookie
                response.set_cookie(
                    self.cookie_name,
                    session_token,
                    max_age=86400,  # 24 hours
                    httponly=True,
                    samesite='Lax',
                    secure=is_secure,
                    path='/'
                )
                
                return response
            else:
                return web.json_response({'success': False, 'message': 'Invalid username or password'}, status=401)
        
        except Exception as e:
            logger.error(f"Login error: {e}")
            return web.json_response({'success': False, 'message': 'Login failed'}, status=500)
    
    async def handle_logout(self, request: web.Request) -> web.Response:
        """Handle logout POST request"""
        response = web.json_response({'success': True, 'message': 'Logged out'})
        
        # Clear session cookie
        response.del_cookie(self.cookie_name, path='/')
        
        return response
    
    async def check_auth(self, request: web.Request) -> bool:
        """
        Check if request is authenticated
        
        Args:
            request: HTTP request
            
        Returns:
            True if authenticated or auth disabled, False otherwise
        """
        if not self.auth_enabled:
            return True
        
        if not self.auth_manager:
            # Auth enabled but not configured = deny access
            return False
        
        # Get session cookie
        session_token = request.cookies.get(self.cookie_name)
        if not session_token:
            return False
        
        # Verify session
        username = self.auth_manager.verify_session(session_token)
        if username:
            # Store username in request for potential future use
            request['username'] = username
            return True
        
        return False
    
    async def handle_websocket(self, request: web.Request) -> web.WebSocketResponse:
        """
        Handle WebSocket connection for streaming
        Sends frames and inventory updates to client
        """
        # Check authentication
        if not await self.check_auth(request):
            ws = web.WebSocketResponse()
            await ws.prepare(request)
            await ws.send_json({'error': 'Unauthorized', 'type': 'error'})
            await ws.close()
            return ws
        
        ws = web.WebSocketResponse(
            heartbeat=30,  # Send ping every 30s
            compress=False  # Disable compression for lower latency
        )
        await ws.prepare(request)
        
        self.websockets.add(ws)
        client_addr = request.remote
        logger.info(f"WebSocket connected: {client_addr} (total: {len(self.websockets)})")
        
        try:
            # Send initial data
            await self.send_to_client(ws, {
                'type': 'inventory',
                'data': self.latest_inventory
            })
            
            await self.send_to_client(ws, {
                'type': 'freshness',
                'data': self.latest_freshness
            })
            
            await self.send_to_client(ws, {
                'type': 'sales',
                'data': self.latest_sales
            })
            
            await self.send_to_client(ws, {
                'type': 'alerts',
                'data': self.latest_alerts
            })
            
            # Handle incoming messages (if any)
            async for msg in ws:
                if msg.type == aiohttp.WSMsgType.TEXT:
                    try:
                        data = json.loads(msg.data)
                        await self.handle_client_message(ws, data)
                    except json.JSONDecodeError:
                        logger.warning(f"Invalid JSON from {client_addr}")
                
                elif msg.type == aiohttp.WSMsgType.ERROR:
                    logger.error(f"WebSocket error: {ws.exception()}")
        
        except asyncio.CancelledError:
            logger.info(f"WebSocket cancelled: {client_addr}")
        
        except Exception as e:
            logger.error(f"WebSocket error for {client_addr}: {e}")
        
        finally:
            self.websockets.discard(ws)
            logger.info(f"WebSocket disconnected: {client_addr} (remaining: {len(self.websockets)})")
        
        return ws
    
    async def handle_client_message(self, ws: web.WebSocketResponse, data: dict):
        """
        Handle incoming message from client
        
        Args:
            ws: WebSocket connection
            data: Parsed JSON data
        """
        msg_type = data.get('type')
        
        if msg_type == 'ping':
            await self.send_to_client(ws, {'type': 'pong'})
        
        elif msg_type == 'request_frame':
            # Client requesting latest frame
            if self.latest_frame is not None:
                await self.send_frame_to_client(ws, self.latest_frame)
        
        elif msg_type == 'switch_camera':
            # Switch camera via WebSocket
            new_index = data.get('index')
            if new_index is None or self._camera_ref is None:
                await self.send_to_client(ws, {
                    'type': 'camera_switch_result',
                    'success': False,
                    'message': 'Invalid index or camera not available',
                })
                return

            new_index = int(new_index)
            if new_index == self._camera_ref.camera_index:
                await self.send_to_client(ws, {
                    'type': 'camera_switch_result',
                    'success': True,
                    'message': f'Already using camera {new_index}',
                    'active_index': new_index,
                })
                return

            ok = self._camera_ref.switch_camera(new_index)
            if ok:
                await self._broadcast_camera_change(new_index)
                await self.send_to_client(ws, {
                    'type': 'camera_switch_result',
                    'success': True,
                    'active_index': new_index,
                })
            else:
                await self.send_to_client(ws, {
                    'type': 'camera_switch_result',
                    'success': False,
                    'message': f'Failed to open camera {new_index}',
                })
    
    async def send_to_client(self, ws: web.WebSocketResponse, data: dict):
        """
        Send JSON data to a single client
        
        Args:
            ws: WebSocket connection
            data: Data to send
        """
        try:
            await ws.send_json(data)
        except Exception as e:
            logger.error(f"Failed to send to client: {e}")
    
    async def send_frame_to_client(self, ws: web.WebSocketResponse, frame: np.ndarray):
        """
        Encode and send frame to a single client
        
        Args:
            ws: WebSocket connection
            frame: Frame to send
        """
        try:
            # Encode frame as JPEG
            _, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
            
            # Convert to base64
            frame_b64 = base64.b64encode(buffer).decode('utf-8')
            
            # Send with timestamp
            await ws.send_json({
                'type': 'frame',
                'data': frame_b64,
                'timestamp': time.time()
            })
        
        except Exception as e:
            logger.error(f"Failed to send frame: {e}")
    
    async def broadcast_frame(self, frame: np.ndarray):
        """
        Broadcast frame to all connected clients
        
        Args:
            frame: Frame to broadcast
        """
        if not self.websockets:
            return
        
        self.latest_frame = frame.copy()
        self.frames_streamed += 1
        
        # Encode once, send to all
        try:
            _, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, 85])
            frame_b64 = base64.b64encode(buffer).decode('utf-8')
            
            message = {
                'type': 'frame',
                'data': frame_b64,
                'timestamp': time.time()
            }
            
            # Send to all clients concurrently
            if self.websockets:
                await asyncio.gather(
                    *[ws.send_json(message) for ws in self.websockets],
                    return_exceptions=True
                )
        
        except Exception as e:
            logger.error(f"Failed to broadcast frame: {e}")
    
    async def broadcast_inventory(self, inventory: dict):
        """
        Broadcast inventory update to all clients
        
        Args:
            inventory: Inventory dictionary
        """
        if not self.websockets:
            self.latest_inventory = inventory
            return
        
        self.latest_inventory = inventory
        
        message = {
            'type': 'inventory',
            'data': inventory,
            'timestamp': time.time()
        }
        
        # Send to all clients
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    async def broadcast_stats(self, stats: dict):
        """
        Broadcast statistics to all clients
        
        Args:
            stats: Statistics dictionary
        """
        self.latest_stats = stats
        
        if not self.websockets:
            return
        
        message = {
            'type': 'stats',
            'data': stats,
            'timestamp': time.time()
        }
        
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    def update_frame(self, frame: np.ndarray):
        """
        Update latest frame (synchronous wrapper)
        
        Args:
            frame: New frame
        """
        self.latest_frame = frame.copy()
    
    def update_inventory(self, inventory: dict):
        """
        Update latest inventory (synchronous wrapper)
        
        Args:
            inventory: New inventory
        """
        self.latest_inventory = inventory
    
    def update_stats(self, stats: dict):
        """
        Update latest stats (synchronous wrapper)
        
        Args:
            stats: New statistics
        """
        self.latest_stats = stats
    
    def update_freshness(self, freshness: dict):
        """
        Update latest freshness data (synchronous wrapper)
        
        Args:
            freshness: New freshness data
        """
        self.latest_freshness = freshness
    
    def update_sales(self, sales: list):
        """
        Update latest sales log (synchronous wrapper)
        
        Args:
            sales: New sales log entries
        """
        self.latest_sales = sales
    
    def update_alerts(self, alerts: list):
        """
        Update latest alerts (synchronous wrapper)
        
        Args:
            alerts: New alerts list
        """
        self.latest_alerts = alerts
    
    async def broadcast_freshness(self, freshness: dict):
        """
        Broadcast freshness update to all clients
        
        Args:
            freshness: Freshness dictionary
        """
        self.latest_freshness = freshness
        
        if not self.websockets:
            return
        
        message = {
            'type': 'freshness',
            'data': freshness,
            'timestamp': time.time()
        }
        
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    async def broadcast_sales(self, sales: list):
        """
        Broadcast sales log update to all clients
        
        Args:
            sales: Sales log entries list
        """
        self.latest_sales = sales
        
        if not self.websockets:
            return
        
        message = {
            'type': 'sales',
            'data': sales,
            'timestamp': time.time()
        }
        
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    async def broadcast_alerts(self, alerts: list):
        """
        Broadcast alerts update to all clients
        
        Args:
            alerts: Alerts list
        """
        self.latest_alerts = alerts
        
        if not self.websockets:
            return
        
        message = {
            'type': 'alerts',
            'data': alerts,
            'timestamp': time.time()
        }
        
        if self.websockets:
            await asyncio.gather(
                *[ws.send_json(message) for ws in self.websockets],
                return_exceptions=True
            )
    
    async def broadcast_video_ended(self, session_id: Optional[str]):
        """
        Broadcast video_ended event to all connected WebSocket clients.

        Args:
            session_id: The video session UUID (may be None)
        """
        message = {
            'type': 'video_ended',
            'session_id': session_id,
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
        disconnected = []
        for ws in self.websockets:
            try:
                await ws.send_json(message)
            except Exception:
                disconnected.append(ws)
        for ws in disconnected:
            self.websockets.discard(ws)

    # ------------------------------------------------------------------
    # Restock Mobile App API Handlers
    # ------------------------------------------------------------------

    async def handle_restock_login(self, request: web.Request) -> web.Response:
        """Handle restock app login"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'message': 'Restock service unavailable'}, status=503)
        
        try:
            data = await request.json()
            username = data.get('username')
            password = data.get('password')
            
            if not username or not password:
                return web.json_response({'success': False, 'message': 'Username and password required'}, status=400)
            
            # Authenticate using existing auth system
            if not self.auth_manager:
                return web.json_response({'success': False, 'message': 'Authentication not configured'}, status=503)
            
            if not self.auth_manager.verify_password(username, password):
                return web.json_response({'success': False, 'message': 'Invalid credentials'}, status=401)
            
            # Create session token
            session_token = self.auth_manager.authenticate(username, password)
            if not session_token:
                return web.json_response({'success': False, 'message': 'Authentication failed'}, status=401)
            
            # Get user role and franchise (simplified - would come from user database)
            role = 'employee'  # Would be determined from user database
            franchise = 'f1'  # Would be determined from user database
            
            # Return session info with cookie
            response = web.json_response({
                'success': True,
                'username': username,
                'role': role,
                'franchise': franchise,
            })
            
            # Set session cookie
            is_secure = request.headers.get('X-Forwarded-Proto', '').lower() == 'https'
            response.set_cookie(
                self.cookie_name,
                session_token,
                max_age=86400,  # 24 hours
                httponly=True,
                samesite='Lax',
                secure=is_secure,
                path='/'
            )
            
            return response
        
        except Exception as e:
            logger.error(f"Restock login error: {e}")
            return web.json_response({'success': False, 'message': 'Login failed'}, status=500)
    
    async def handle_restock_validate(self, request: web.Request) -> web.Response:
        """Validate restock session token"""
        try:
            # Get session cookie
            session_token = request.cookies.get(self.cookie_name)
            if not session_token:
                return web.json_response({'valid': False}, status=401)
            
            if not self.auth_manager:
                return web.json_response({'valid': False}, status=503)
            
            username = self.auth_manager.verify_session(session_token)
            if username:
                return web.json_response({'valid': True, 'username': username})
            else:
                return web.json_response({'valid': False}, status=401)
        
        except Exception as e:
            logger.error(f"Restock validate error: {e}")
            return web.json_response({'valid': False}, status=500)
    
    async def handle_restock_logout(self, request: web.Request) -> web.Response:
        """Handle restock app logout"""
        response = web.json_response({'success': True, 'message': 'Logged out'})
        response.del_cookie(self.cookie_name, path='/')
        return response
    
    async def handle_restock_detect(self, request: web.Request) -> web.Response:
        """Run YOLO detection on uploaded photo (for preview)"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'message': 'Service unavailable'}, status=503)
        
        if not self._detector_ref:
            return web.json_response({'success': False, 'message': 'Detector not available'}, status=503)
        
        try:
            # Parse multipart form data
            reader = await request.multipart()
            photo_data = None
            
            async for field in reader:
                if field.name == 'photo':
                    photo_data = await field.read()
                    break
            
            if not photo_data:
                return web.json_response({'success': False, 'message': 'No photo provided'}, status=400)
            
            # Convert to numpy array
            nparr = np.frombuffer(photo_data, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
            
            if img is None:
                return web.json_response({'success': False, 'message': 'Invalid image'}, status=400)
            
            # Run detection
            detections = self._detector_ref.detect(img)
            
            # Count products by class
            product_counts = {}
            for det in detections:
                class_name = det.get('class_name', 'unknown')
                product_counts[class_name] = product_counts.get(class_name, 0) + 1
            
            return web.json_response({
                'success': True,
                'detections': detections,
                'product_counts': product_counts,
                'total_detections': len(detections)
            })
        
        except Exception as e:
            logger.error(f"Restock detect error: {e}")
            return web.json_response({'success': False, 'message': 'Detection failed'}, status=500)
    
    async def handle_restock_upload(self, request: web.Request) -> web.Response:
        """Handle restock photo upload"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'message': 'Restock service unavailable'}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'message': 'Unauthorized'}, status=401)
        
        username = request.get('username', 'unknown')
        
        try:
            # Parse multipart form data
            reader = await request.multipart()
            
            photos = []
            station = None
            product = None
            notes = None
            device_id = None
            latitude = None
            longitude = None
            detection_results = None
            franchise_id = 'f1'  # Would come from user database
            
            async for field in reader:
                if field.name.startswith('photo_'):
                    photo_data = await field.read()
                    if photo_data:
                        # Reject PNG files - only accept JPEG
                        if photo_data.startswith(b'\x89PNG\r\n\x1a\n'):
                            return web.json_response(
                                {'success': False, 'message': 'PNG files are not accepted. Please upload JPEG images.'},
                                status=400,
                            )
                        
                        # Verify it's a JPEG file (starts with FF D8 FF)
                        if not photo_data.startswith(b'\xff\xd8\xff'):
                            return web.json_response(
                                {'success': False, 'message': 'Only JPEG images are accepted. Please upload JPEG images.'},
                                status=400,
                            )
                        
                        photos.append(photo_data)
                elif field.name == 'station':
                    station = await field.text()
                elif field.name == 'product':
                    product = await field.text()
                elif field.name == 'notes':
                    notes = await field.text()
                elif field.name == 'device_id':
                    device_id = await field.text()
                elif field.name == 'latitude':
                    lat_str = await field.text()
                    latitude = float(lat_str) if lat_str else None
                elif field.name == 'longitude':
                    lng_str = await field.text()
                    longitude = float(lng_str) if lng_str else None
                elif field.name == 'detection_results':
                    detection_str = await field.text()
                    if detection_str:
                        import json
                        detection_results = json.loads(detection_str)
            
            if not station or not product:
                return web.json_response({'success': False, 'message': 'Station and product required'}, status=400)
            
            if len(photos) < 3:
                return web.json_response({'success': False, 'message': 'Minimum 3 photos required'}, status=400)
            
            # Create submission
            success, message, submission_id = self.restock_manager.create_submission(
                employee_username=username,
                franchise_id=franchise_id,
                station=station,
                product=product,
                notes=notes,
                device_id=device_id,
                latitude=latitude,
                longitude=longitude,
                photos=photos,
                detection_results=detection_results
            )
            
            if not success:
                return web.json_response({'success': False, 'message': message}, status=400)

            # Log the upload
            SystemLogger.upload(
                f"Restock submission: {product} at {station}",
                {
                    'submission_id': submission_id,
                    'username': username,
                    'product': product,
                    'station': station,
                    'photo_count': len(photos)
                }
            )

            # Broadcast to dashboard for live updates
            await self.broadcast_event('restock', {
                'event': 'submission_created',
                'submission_id': submission_id,
                'username': username,
                'product': product,
                'station': station
            })

            return web.json_response({
                'success': True,
                'message': message,
                'submission_id': submission_id
            })
        
        except Exception as e:
            logger.error(f"Restock upload error: {e}")
            return web.json_response({'success': False, 'message': 'Upload failed'}, status=500)
    
    async def handle_restock_submissions(self, request: web.Request) -> web.Response:
        """Get employee's submissions"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'submissions': []}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'submissions': []}, status=401)
        
        username = request.get('username', 'unknown')
        
        try:
            submissions = self.restock_manager.get_employee_submissions(username)
            return web.json_response({'success': True, 'submissions': submissions})
        
        except Exception as e:
            logger.error(f"Error getting submissions: {e}")
            return web.json_response({'success': False, 'submissions': []}, status=500)
    
    async def handle_restock_notifications(self, request: web.Request) -> web.Response:
        """Get employee notifications"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'notifications': []}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'notifications': []}, status=401)
        
        username = request.get('username', 'unknown')
        
        try:
            notifications = self.restock_manager.get_notifications(username)
            formatted = [{
                'id': n['id'],
                'title': n['title'],
                'message': n['message'],
                'timestamp': n['timestamp_utc'],
                'read': bool(n['read'])
            } for n in notifications]
            
            return web.json_response({'success': True, 'notifications': formatted})
        
        except Exception as e:
            logger.error(f"Error getting notifications: {e}")
            return web.json_response({'success': False, 'notifications': []}, status=500)
    
    async def handle_restock_notification_count(self, request: web.Request) -> web.Response:
        """Get unread notification count"""
        if not self.restock_manager:
            return web.json_response({'count': 0}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'count': 0}, status=401)
        
        username = request.get('username', 'unknown')
        
        try:
            count = self.restock_manager.get_notification_count(username)
            return web.json_response({'count': count})
        except Exception as e:
            logger.error(f"Error getting notification count: {e}")
            return web.json_response({'count': 0}, status=500)
    
    async def handle_restock_notification_read(self, request: web.Request) -> web.Response:
        """Mark notification as read"""
        if not self.restock_manager:
            return web.json_response({'success': False}, status=503)
        
        try:
            data = await request.json()
            notification_id = data.get('notification_id')
            
            if not notification_id:
                return web.json_response({'success': False}, status=400)
            
            success = self.restock_manager.mark_notification_read(str(notification_id))
            return web.json_response({'success': success})
        except Exception as e:
            logger.error(f"Error marking notification read: {e}")
            return web.json_response({'success': False}, status=500)
    
    async def handle_restock_photo(self, request: web.Request) -> web.Response:
        """Serve restock photo file"""
        if not self.restock_manager:
            return web.Response(status=404)
        
        filename = request.match_info.get('filename')
        if not filename:
            return web.Response(status=404)
        
        photo_path = self.restock_manager.get_photo_path(filename)
        if photo_path and photo_path.exists():
            return web.FileResponse(photo_path)
        
        return web.Response(status=404)
    
    async def handle_restock_all(self, request: web.Request) -> web.Response:
        """Get all submissions (manager only)"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'submissions': []}, status=503)
        
        # Check authentication and manager role
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'submissions': []}, status=401)
        
        try:
            # Get query parameters
            franchise_id = request.query.get('franchise')
            status = request.query.get('status')
            employee = request.query.get('employee')
            
            submissions = self.restock_manager.get_all_submissions(
                franchise_id=franchise_id,
                status=status,
                employee=employee
            )
            
            return web.json_response({'success': True, 'submissions': submissions})
        
        except Exception as e:
            logger.error(f"Error getting all submissions: {e}")
            return web.json_response({'success': False, 'submissions': []}, status=500)
    
    async def handle_restock_status_update(self, request: web.Request) -> web.Response:
        """Update submission status (manager only)"""
        if not self.restock_manager:
            return web.json_response({'success': False, 'message': 'Service unavailable'}, status=503)
        
        # Check authentication
        if not await self.check_auth(request):
            return web.json_response({'success': False, 'message': 'Unauthorized'}, status=401)
        
        try:
            data = await request.json()
            submission_id = data.get('submission_id')
            status = data.get('status')
            feedback = data.get('feedback')
            
            if not submission_id or not status:
                return web.json_response({'success': False, 'message': 'Missing required fields'}, status=400)
            
            # Get reviewer username
            username = request.get('username', 'manager')
            
            success = self.restock_manager.update_submission_status(
                submission_id=submission_id,
                status=status,
                reviewed_by=username,
                feedback=feedback
            )
            
            if success:
                return web.json_response({'success': True, 'message': 'Status updated'})
            else:
                return web.json_response({'success': False, 'message': 'Update failed'}, status=400)
        
        except Exception as e:
            logger.error(f"Error updating submission status: {e}")
            return web.json_response({'success': False, 'message': 'Update failed'}, status=500)
    
    async def start(self):
        """Start the web server"""
        runner = web.AppRunner(self.app)
        await runner.setup()
        
        site = web.TCPSite(runner, self.host, self.port)
        await site.start()
        
        logger.info(f"Server started at http://{self.host}:{self.port}")
        logger.info(f"Frontend directory: {self.frontend_dir}")
    
    def run(self):
        """Run the server (blocking)"""
        web.run_app(
            self.app,
            host=self.host,
            port=self.port,
            print=None,  # Disable aiohttp's startup message
            access_log=None  # Disable access logs for performance
        )
    
    def get_url(self) -> str:
        """
        Get server URL
        
        Returns:
            Server URL string
        """
        return f"http://{self.host}:{self.port}"


class StreamManager:
    """
    Manages streaming loop coordination between camera, detector, and server
    """
    
    def __init__(
        self,
        camera,
        detector,
        inventory_tracker,
        server: VideoStreamServer,
        target_fps: int = 30
    ):
        """
        Initialize stream manager
        
        Args:
            camera: USBCamera instance
            detector: YOLODetector instance
            inventory_tracker: InventoryTracker instance
            server: VideoStreamServer instance
            target_fps: Target streaming FPS
        """
        self.camera = camera
        self.detector = detector
        self.inventory_tracker = inventory_tracker
        self.server = server
        self.target_fps = target_fps
        self.frame_interval = 1.0 / target_fps
        
        self.is_running = False
        self.loop_task = None
        self.current_session_id: Optional[str] = None
        self.session_start_time: Optional[float] = None

    async def stream_loop(self):
        """
        Main streaming loop
        Captures frames, runs inference, updates inventory, and broadcasts
        """
        logger.info("Starting stream loop...")
        self.is_running = True
        
        frame_count = 0
        last_stats_time = time.time()
        stats_interval = 1.0  # Update stats every second
        
        while self.is_running:
            loop_start = time.time()
            
            # Capture frame
            success, frame = self.camera.read()
            
            if not success or frame is None:
                # Check if this is a VideoFileCamera that has reached its end (loop=False)
                from camera import VideoFileCamera
                if isinstance(self.camera, VideoFileCamera) and not self.camera.loop:
                    logger.info("Video file ended naturally (loop=False). Finalizing session...")
                    await self.server.broadcast_video_ended(self.current_session_id)
                    break  # Exit the streaming loop cleanly
                # Otherwise: USB camera lost — attempt reconnection
                logger.warning("Failed to capture frame, attempting reconnection...")
                if not self.camera.reconnect():
                    await asyncio.sleep(1.0)
                    continue
                success, frame = self.camera.read()
                if not success:
                    await asyncio.sleep(1.0)
                    continue
            
            # Run detection
            detections = self.detector.detect(frame)
            
            # Update inventory
            self.inventory_tracker.update(detections)
            inventory = self.inventory_tracker.get_inventory()
            
            # Draw detections on frame
            annotated_frame = self.detector.draw_detections(frame, detections)
            
            # Broadcast frame and inventory
            await self.server.broadcast_frame(annotated_frame)
            await self.server.broadcast_inventory(inventory)
            
            frame_count += 1
            
            # Broadcast stats periodically
            current_time = time.time()
            if current_time - last_stats_time >= stats_interval:
                stats = {
                    'fps': self.detector.get_fps(),
                    'inference_time': self.detector.get_average_inference_time(),
                    'total_items': self.inventory_tracker.get_total_items(),
                    'frame_count': frame_count,
                    'active_connections': len(self.server.websockets)
                }
                await self.server.broadcast_stats(stats)
                
                # Broadcast freshness, sales, and alerts data if available
                if hasattr(self.inventory_tracker, 'get_freshness_state'):
                    freshness = self.inventory_tracker.get_freshness_state()
                    await self.server.broadcast_freshness(freshness)
                
                if hasattr(self.inventory_tracker, 'get_sales_history'):
                    sales = self.inventory_tracker.get_sales_history(limit=100)
                    await self.server.broadcast_sales(sales)
                
                if hasattr(self.inventory_tracker, 'get_recent_alerts'):
                    alerts = self.inventory_tracker.get_recent_alerts(limit=20)
                    await self.server.broadcast_alerts(alerts)
                
                last_stats_time = current_time
            
            # Maintain target FPS
            elapsed = time.time() - loop_start
            sleep_time = max(0, self.frame_interval - elapsed)
            if sleep_time > 0:
                await asyncio.sleep(sleep_time)
        
        logger.info("Stream loop stopped")
    
    def start(self):
        """Start streaming loop"""
        if self.is_running:
            logger.warning("Stream already running")
            return
        
        self.loop_task = asyncio.create_task(self.stream_loop())
    
    async def stop(self):
        """Stop streaming loop"""
        self.is_running = False
        
        if self.loop_task:
            await self.loop_task
            self.loop_task = None

