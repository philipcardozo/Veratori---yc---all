"""
Analytics Module for Sales Data Processing
Handles CSV/XLSX uploads, validation, aggregation, and reporting
"""

import csv
import io
import json
import logging
import os
import sqlite3
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import uuid

logger = logging.getLogger(__name__)

# Required and optional fields for sales data
REQUIRED_FIELDS = {'product', 'quantity', 'price', 'timestamp'}
OPTIONAL_FIELDS = {'location', 'employee_id', 'cost', 'category'}
ALL_FIELDS = REQUIRED_FIELDS | OPTIONAL_FIELDS


class AnalyticsProcessor:
    """
    Processes uploaded sales data files (CSV/XLSX)
    Validates schema, aggregates metrics, and stores results
    """

    def __init__(self, db_path: Optional[Path] = None, config: Optional[dict] = None):
        """
        Initialize analytics processor.

        Args:
            db_path: Path to SQLite database
            config: Configuration dictionary with thresholds
        """
        self.db_path = db_path or Path(__file__).parent.parent / 'data' / 'inventory.db'
        self.export_dir = Path(__file__).parent.parent / 'run' / 'exports'
        self.export_dir.mkdir(parents=True, exist_ok=True)

        # Load config for thresholds
        self.config = config or self._load_config()
        self.low_stock_thresholds = self.config.get('alerts', {}).get('low_stock_thresholds', {})

        # Default threshold if product not in config
        self.default_threshold = 5

        # Ensure database tables exist
        self._init_database()

    def _load_config(self) -> dict:
        """Load configuration from config.yaml."""
        try:
            import yaml
            config_path = Path(__file__).parent.parent / 'config' / 'config.yaml'
            if config_path.exists():
                with open(config_path, 'r') as f:
                    return yaml.safe_load(f) or {}
        except Exception as e:
            logger.warning(f"Could not load config: {e}")
        return {}

    def _init_database(self):
        """Initialize database tables for analytics data."""
        self.db_path.parent.mkdir(parents=True, exist_ok=True)

        with sqlite3.connect(str(self.db_path)) as conn:
            cursor = conn.cursor()

            # Sales uploads table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS sales_uploads (
                    id TEXT PRIMARY KEY,
                    filename TEXT NOT NULL,
                    upload_time_utc REAL NOT NULL,
                    total_records INTEGER NOT NULL,
                    valid_records INTEGER NOT NULL,
                    total_revenue REAL NOT NULL,
                    total_units INTEGER NOT NULL,
                    export_path TEXT,
                    status TEXT DEFAULT 'completed'
                )
            """)

            # Processed sales records table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS processed_sales (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    upload_id TEXT NOT NULL,
                    product TEXT NOT NULL,
                    quantity INTEGER NOT NULL,
                    price REAL NOT NULL,
                    revenue REAL NOT NULL,
                    timestamp_utc REAL NOT NULL,
                    timestamp_est TEXT NOT NULL,
                    date_est TEXT NOT NULL,
                    location TEXT,
                    employee_id TEXT,
                    cost REAL,
                    category TEXT,
                    FOREIGN KEY (upload_id) REFERENCES sales_uploads(id)
                )
            """)

            # Analytics summaries table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS analytics_summaries (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    upload_id TEXT NOT NULL,
                    summary_type TEXT NOT NULL,
                    summary_key TEXT NOT NULL,
                    summary_value REAL NOT NULL,
                    metadata TEXT,
                    FOREIGN KEY (upload_id) REFERENCES sales_uploads(id)
                )
            """)

            # Create indexes
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_processed_sales_upload ON processed_sales(upload_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_processed_sales_product ON processed_sales(product)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_processed_sales_date ON processed_sales(date_est)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_analytics_summaries_upload ON analytics_summaries(upload_id)")

            conn.commit()

    def validate_schema(self, headers: List[str]) -> Tuple[bool, str, List[str]]:
        """
        Validate that required headers are present.

        Args:
            headers: List of column headers from the file

        Returns:
            Tuple of (is_valid, error_message, normalized_headers)
        """
        # Normalize headers (lowercase, strip whitespace)
        normalized = [h.lower().strip().replace(' ', '_') for h in headers]

        # Check for required fields
        missing = REQUIRED_FIELDS - set(normalized)
        if missing:
            return False, f"Missing required columns: {', '.join(sorted(missing))}", []

        # Filter to only recognized fields
        valid_headers = [h for h in normalized if h in ALL_FIELDS]

        return True, "", valid_headers

    def parse_csv(self, content: bytes, filename: str) -> Tuple[bool, str, List[dict]]:
        """
        Parse CSV file content.

        Args:
            content: Raw file bytes
            filename: Original filename

        Returns:
            Tuple of (success, error_message, records)
        """
        try:
            # Try different encodings
            text = None
            for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if text is None:
                return False, "Could not decode file with supported encodings", []

            # Parse CSV
            reader = csv.DictReader(io.StringIO(text))
            headers = reader.fieldnames or []

            if not headers:
                return False, "CSV file has no headers", []

            # Validate schema
            is_valid, error, valid_headers = self.validate_schema(headers)
            if not is_valid:
                return False, error, []

            # Parse records
            records = []
            for row_num, row in enumerate(reader, start=2):
                try:
                    record = self._parse_row(row, headers)
                    if record:
                        records.append(record)
                except Exception as e:
                    logger.warning(f"Error parsing row {row_num}: {e}")
                    continue

            if not records:
                return False, "No valid records found in file", []

            return True, "", records

        except Exception as e:
            logger.error(f"CSV parsing error: {e}")
            return False, f"Failed to parse CSV: {str(e)}", []

    def parse_xlsx(self, content: bytes, filename: str) -> Tuple[bool, str, List[dict]]:
        """
        Parse XLSX file content.

        Args:
            content: Raw file bytes
            filename: Original filename

        Returns:
            Tuple of (success, error_message, records)
        """
        try:
            from openpyxl import load_workbook

            # Load workbook from bytes
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active

            if ws is None:
                return False, "Excel file has no active worksheet", []

            # Get headers from first row
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return False, "Excel file is empty", []

            headers = [str(h) if h else '' for h in rows[0]]

            # Validate schema
            is_valid, error, valid_headers = self.validate_schema(headers)
            if not is_valid:
                return False, error, []

            # Parse records
            records = []
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    row_dict = dict(zip(headers, row))
                    record = self._parse_row(row_dict, headers)
                    if record:
                        records.append(record)
                except Exception as e:
                    logger.warning(f"Error parsing row {row_num}: {e}")
                    continue

            wb.close()

            if not records:
                return False, "No valid records found in file", []

            return True, "", records

        except ImportError:
            return False, "openpyxl not installed - cannot process Excel files", []
        except Exception as e:
            logger.error(f"XLSX parsing error: {e}")
            return False, f"Failed to parse Excel file: {str(e)}", []

    def _parse_row(self, row: dict, headers: List[str]) -> Optional[dict]:
        """
        Parse and validate a single row.

        Args:
            row: Row data as dictionary
            headers: Original headers

        Returns:
            Parsed record or None if invalid
        """
        # Normalize keys
        normalized_row = {}
        for key, value in row.items():
            norm_key = key.lower().strip().replace(' ', '_') if key else ''
            if norm_key in ALL_FIELDS:
                normalized_row[norm_key] = value

        # Validate required fields
        for field in REQUIRED_FIELDS:
            if field not in normalized_row or normalized_row[field] is None:
                return None
            if str(normalized_row[field]).strip() == '':
                return None

        # Parse and validate values
        try:
            record = {
                'product': str(normalized_row['product']).strip(),
                'quantity': int(float(str(normalized_row['quantity']).replace(',', ''))),
                'price': float(str(normalized_row['price']).replace('$', '').replace(',', '')),
            }

            # Validate positive values
            if record['quantity'] <= 0 or record['price'] < 0:
                return None

            # Parse timestamp
            ts_value = normalized_row['timestamp']
            record['timestamp_utc'] = self._parse_timestamp(ts_value)
            if record['timestamp_utc'] is None:
                return None

            # Calculate revenue
            record['revenue'] = record['quantity'] * record['price']

            # Convert to EST for display
            utc_dt = datetime.fromtimestamp(record['timestamp_utc'], tz=timezone.utc)
            est_offset = timedelta(hours=-5)  # EST
            est_dt = utc_dt + est_offset
            record['timestamp_est'] = est_dt.strftime('%Y-%m-%d %H:%M:%S')
            record['date_est'] = est_dt.strftime('%Y-%m-%d')

            # Optional fields
            record['location'] = str(normalized_row.get('location', '')).strip() or None
            record['employee_id'] = str(normalized_row.get('employee_id', '')).strip() or None
            record['category'] = str(normalized_row.get('category', '')).strip() or None

            cost_val = normalized_row.get('cost')
            if cost_val is not None and str(cost_val).strip():
                try:
                    record['cost'] = float(str(cost_val).replace('$', '').replace(',', ''))
                except:
                    record['cost'] = None
            else:
                record['cost'] = None

            return record

        except Exception as e:
            logger.debug(f"Row parsing error: {e}")
            return None

    def _parse_timestamp(self, value: Any) -> Optional[float]:
        """
        Parse various timestamp formats.

        Args:
            value: Timestamp value (string, datetime, or number)

        Returns:
            Unix timestamp or None
        """
        if value is None:
            return None

        # Already a datetime
        if isinstance(value, datetime):
            if value.tzinfo is None:
                value = value.replace(tzinfo=timezone.utc)
            return value.timestamp()

        # Number (unix timestamp)
        if isinstance(value, (int, float)):
            return float(value)

        # String parsing
        ts_str = str(value).strip()

        # Common formats to try
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%dT%H:%M:%SZ',
            '%Y-%m-%dT%H:%M:%S.%f',
            '%Y-%m-%dT%H:%M:%S.%fZ',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y %H:%M',
            '%m/%d/%Y',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y',
            '%Y-%m-%d',
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(ts_str, fmt)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt.timestamp()
            except ValueError:
                continue

        # Try ISO format
        try:
            dt = datetime.fromisoformat(ts_str.replace('Z', '+00:00'))
            return dt.timestamp()
        except:
            pass

        return None

    def process_file(self, content: bytes, filename: str) -> dict:
        """
        Process uploaded sales file.

        Args:
            content: Raw file bytes
            filename: Original filename

        Returns:
            Processing result with analytics data
        """
        upload_id = str(uuid.uuid4())[:12]
        upload_time = datetime.now(timezone.utc).timestamp()

        # Determine file type and parse
        ext = Path(filename).suffix.lower()

        if ext == '.csv':
            success, error, records = self.parse_csv(content, filename)
        elif ext in ['.xlsx', '.xls']:
            success, error, records = self.parse_xlsx(content, filename)
        else:
            return {
                'success': False,
                'error': f'Unsupported file type: {ext}. Use CSV or XLSX.',
                'upload_id': upload_id
            }

        if not success:
            return {
                'success': False,
                'error': error,
                'upload_id': upload_id
            }

        # Calculate aggregations
        aggregations = self._calculate_aggregations(records)

        # Generate insights
        insights = self._generate_insights(records, aggregations)

        # Save to database
        export_path = self._save_results(upload_id, filename, upload_time, records, aggregations)

        # Build response
        return {
            'success': True,
            'upload_id': upload_id,
            'filename': filename,
            'upload_time': datetime.fromtimestamp(upload_time, tz=timezone.utc).isoformat(),
            'summary': {
                'total_records': len(records),
                'total_revenue': aggregations['total_revenue'],
                'total_units': aggregations['total_units'],
                'active_products': len(aggregations['revenue_by_product']),
                'date_range': {
                    'start': aggregations['date_range']['start'],
                    'end': aggregations['date_range']['end']
                }
            },
            'revenue_by_product': aggregations['revenue_by_product'],
            'units_by_product': aggregations['units_by_product'],
            'sales_by_date': aggregations['sales_by_date'],
            'revenue_by_date': aggregations['revenue_by_date'],
            'low_stock_alerts': aggregations['low_stock_alerts'],
            'insights': insights,
            'export_path': str(export_path) if export_path else None
        }

    def _calculate_aggregations(self, records: List[dict]) -> dict:
        """
        Calculate all aggregation metrics.

        Args:
            records: List of parsed sales records

        Returns:
            Aggregation results
        """
        total_revenue = 0.0
        total_units = 0
        revenue_by_product = {}
        units_by_product = {}
        sales_by_date = {}
        revenue_by_date = {}
        dates = []

        for record in records:
            product = record['product']
            quantity = record['quantity']
            revenue = record['revenue']
            date = record['date_est']

            # Totals
            total_revenue += revenue
            total_units += quantity

            # By product
            revenue_by_product[product] = revenue_by_product.get(product, 0) + revenue
            units_by_product[product] = units_by_product.get(product, 0) + quantity

            # By date
            sales_by_date[date] = sales_by_date.get(date, 0) + quantity
            revenue_by_date[date] = revenue_by_date.get(date, 0) + revenue

            dates.append(date)

        # Sort by date
        sorted_dates = sorted(set(dates))
        sales_by_date_sorted = [{'date': d, 'units': sales_by_date[d]} for d in sorted_dates]
        revenue_by_date_sorted = [{'date': d, 'revenue': revenue_by_date[d]} for d in sorted_dates]

        # Sort products by revenue
        revenue_by_product_sorted = [
            {'product': k, 'revenue': v}
            for k, v in sorted(revenue_by_product.items(), key=lambda x: -x[1])
        ]
        units_by_product_sorted = [
            {'product': k, 'units': v}
            for k, v in sorted(units_by_product.items(), key=lambda x: -x[1])
        ]

        # Check low stock
        low_stock_alerts = []
        for product, units in units_by_product.items():
            threshold = self.low_stock_thresholds.get(
                product.lower(),
                self.default_threshold
            )
            # Check if remaining inventory might be low based on sales velocity
            avg_daily_sales = units / max(len(sorted_dates), 1)
            if avg_daily_sales > 0:
                days_of_stock = threshold / avg_daily_sales
                if days_of_stock < 3:  # Less than 3 days of stock
                    low_stock_alerts.append({
                        'product': product,
                        'threshold': threshold,
                        'avg_daily_sales': round(avg_daily_sales, 1),
                        'risk_level': 'high' if days_of_stock < 1 else 'medium'
                    })

        return {
            'total_revenue': round(total_revenue, 2),
            'total_units': total_units,
            'revenue_by_product': revenue_by_product_sorted,
            'units_by_product': units_by_product_sorted,
            'sales_by_date': sales_by_date_sorted,
            'revenue_by_date': revenue_by_date_sorted,
            'low_stock_alerts': low_stock_alerts,
            'date_range': {
                'start': sorted_dates[0] if sorted_dates else None,
                'end': sorted_dates[-1] if sorted_dates else None
            }
        }

    def _generate_insights(self, records: List[dict], aggregations: dict) -> List[dict]:
        """
        Generate automated insights from the data.

        Args:
            records: Parsed sales records
            aggregations: Calculated aggregations

        Returns:
            List of insight objects
        """
        insights = []

        # Best Selling Product
        if aggregations['units_by_product']:
            best_product = aggregations['units_by_product'][0]
            insights.append({
                'type': 'best_seller',
                'title': 'Best Selling Product',
                'value': best_product['product'],
                'detail': f"{best_product['units']:,} units sold",
                'icon': 'trophy'
            })

        # Highest Revenue Day
        if aggregations['revenue_by_date']:
            best_day = max(aggregations['revenue_by_date'], key=lambda x: x['revenue'])
            insights.append({
                'type': 'top_day',
                'title': 'Highest Revenue Day',
                'value': best_day['date'],
                'detail': f"${best_day['revenue']:,.2f} revenue",
                'icon': 'calendar'
            })

        # Inventory Risk Alert
        if aggregations['low_stock_alerts']:
            high_risk = [a for a in aggregations['low_stock_alerts'] if a['risk_level'] == 'high']
            if high_risk:
                insights.append({
                    'type': 'inventory_risk',
                    'title': 'Inventory Risk Alert',
                    'value': f"{len(high_risk)} product(s)",
                    'detail': f"High velocity items: {', '.join(a['product'] for a in high_risk[:3])}",
                    'icon': 'alert',
                    'severity': 'warning'
                })
            elif aggregations['low_stock_alerts']:
                insights.append({
                    'type': 'inventory_risk',
                    'title': 'Inventory Monitor',
                    'value': f"{len(aggregations['low_stock_alerts'])} product(s)",
                    'detail': 'Items approaching restock threshold',
                    'icon': 'alert',
                    'severity': 'info'
                })

        # Top Revenue Product
        if aggregations['revenue_by_product']:
            top_revenue = aggregations['revenue_by_product'][0]
            insights.append({
                'type': 'top_revenue',
                'title': 'Top Revenue Product',
                'value': top_revenue['product'],
                'detail': f"${top_revenue['revenue']:,.2f} total revenue",
                'icon': 'dollar'
            })

        return insights

    def _save_results(self, upload_id: str, filename: str, upload_time: float,
                      records: List[dict], aggregations: dict) -> Optional[Path]:
        """
        Save processing results to database and export CSV.

        Args:
            upload_id: Unique upload identifier
            filename: Original filename
            upload_time: Upload timestamp
            records: Parsed records
            aggregations: Calculated aggregations

        Returns:
            Path to exported CSV or None
        """
        export_path = None

        try:
            with sqlite3.connect(str(self.db_path)) as conn:
                cursor = conn.cursor()

                # Save upload record
                cursor.execute("""
                    INSERT INTO sales_uploads
                    (id, filename, upload_time_utc, total_records, valid_records,
                     total_revenue, total_units, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    upload_id,
                    filename,
                    upload_time,
                    len(records),
                    len(records),
                    aggregations['total_revenue'],
                    aggregations['total_units'],
                    'completed'
                ))

                # Save processed sales records
                for record in records:
                    cursor.execute("""
                        INSERT INTO processed_sales
                        (upload_id, product, quantity, price, revenue,
                         timestamp_utc, timestamp_est, date_est,
                         location, employee_id, cost, category)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        upload_id,
                        record['product'],
                        record['quantity'],
                        record['price'],
                        record['revenue'],
                        record['timestamp_utc'],
                        record['timestamp_est'],
                        record['date_est'],
                        record.get('location'),
                        record.get('employee_id'),
                        record.get('cost'),
                        record.get('category')
                    ))

                # Save analytics summaries
                for item in aggregations['revenue_by_product']:
                    cursor.execute("""
                        INSERT INTO analytics_summaries
                        (upload_id, summary_type, summary_key, summary_value)
                        VALUES (?, ?, ?, ?)
                    """, (upload_id, 'revenue_by_product', item['product'], item['revenue']))

                for item in aggregations['revenue_by_date']:
                    cursor.execute("""
                        INSERT INTO analytics_summaries
                        (upload_id, summary_type, summary_key, summary_value)
                        VALUES (?, ?, ?, ?)
                    """, (upload_id, 'revenue_by_date', item['date'], item['revenue']))

                conn.commit()

            # Export cleaned CSV
            export_path = self._export_csv(upload_id, records, aggregations)

            # Update export path in database
            if export_path:
                with sqlite3.connect(str(self.db_path)) as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        "UPDATE sales_uploads SET export_path = ? WHERE id = ?",
                        (str(export_path), upload_id)
                    )
                    conn.commit()

        except Exception as e:
            logger.error(f"Error saving results: {e}", exc_info=True)

        return export_path

    def _export_csv(self, upload_id: str, records: List[dict], aggregations: dict) -> Optional[Path]:
        """
        Export cleaned data to CSV file.

        Args:
            upload_id: Upload identifier
            records: Parsed records
            aggregations: Calculated aggregations

        Returns:
            Path to exported file or None
        """
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"sales_export_{upload_id}_{timestamp}.csv"
            export_path = self.export_dir / filename

            with open(export_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Write header
                writer.writerow([
                    'Product', 'Quantity', 'Price', 'Revenue',
                    'Timestamp (EST)', 'Date (EST)',
                    'Location', 'Employee ID', 'Cost', 'Category'
                ])

                # Write records
                for record in records:
                    writer.writerow([
                        record['product'],
                        record['quantity'],
                        f"${record['price']:.2f}",
                        f"${record['revenue']:.2f}",
                        record['timestamp_est'],
                        record['date_est'],
                        record.get('location', ''),
                        record.get('employee_id', ''),
                        f"${record['cost']:.2f}" if record.get('cost') else '',
                        record.get('category', '')
                    ])

                # Write summary section
                writer.writerow([])
                writer.writerow(['--- SUMMARY ---'])
                writer.writerow(['Total Revenue', f"${aggregations['total_revenue']:,.2f}"])
                writer.writerow(['Total Units', aggregations['total_units']])
                writer.writerow(['Active Products', len(aggregations['revenue_by_product'])])

            return export_path

        except Exception as e:
            logger.error(f"Error exporting CSV: {e}")
            return None

    def get_latest_analytics(self) -> Optional[dict]:
        """
        Get the most recent analytics data.

        Returns:
            Latest analytics data or None
        """
        try:
            with sqlite3.connect(str(self.db_path)) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                # Get latest upload
                cursor.execute("""
                    SELECT * FROM sales_uploads
                    ORDER BY upload_time_utc DESC
                    LIMIT 1
                """)
                upload = cursor.fetchone()

                if not upload:
                    return None

                upload_id = upload['id']

                # Get revenue by product
                cursor.execute("""
                    SELECT summary_key as product, summary_value as revenue
                    FROM analytics_summaries
                    WHERE upload_id = ? AND summary_type = 'revenue_by_product'
                    ORDER BY summary_value DESC
                """, (upload_id,))
                revenue_by_product = [dict(row) for row in cursor.fetchall()]

                # Get units by product
                cursor.execute("""
                    SELECT product, SUM(quantity) as units
                    FROM processed_sales
                    WHERE upload_id = ?
                    GROUP BY product
                    ORDER BY units DESC
                """, (upload_id,))
                units_by_product = [dict(row) for row in cursor.fetchall()]

                # Get revenue by date
                cursor.execute("""
                    SELECT summary_key as date, summary_value as revenue
                    FROM analytics_summaries
                    WHERE upload_id = ? AND summary_type = 'revenue_by_date'
                    ORDER BY summary_key ASC
                """, (upload_id,))
                revenue_by_date = [dict(row) for row in cursor.fetchall()]

                # Get sales by date
                cursor.execute("""
                    SELECT date_est as date, SUM(quantity) as units
                    FROM processed_sales
                    WHERE upload_id = ?
                    GROUP BY date_est
                    ORDER BY date_est ASC
                """, (upload_id,))
                sales_by_date = [dict(row) for row in cursor.fetchall()]

                # Calculate low stock alerts
                low_stock_alerts = []
                for item in units_by_product:
                    product = item['product']
                    units = item['units']
                    threshold = self.low_stock_thresholds.get(
                        product.lower(),
                        self.default_threshold
                    )
                    num_days = len(revenue_by_date) or 1
                    avg_daily = units / num_days
                    if avg_daily > 0:
                        days_of_stock = threshold / avg_daily
                        if days_of_stock < 3:
                            low_stock_alerts.append({
                                'product': product,
                                'threshold': threshold,
                                'avg_daily_sales': round(avg_daily, 1),
                                'risk_level': 'high' if days_of_stock < 1 else 'medium'
                            })

                # Generate insights
                insights = []
                if units_by_product:
                    insights.append({
                        'type': 'best_seller',
                        'title': 'Best Selling Product',
                        'value': units_by_product[0]['product'],
                        'detail': f"{units_by_product[0]['units']:,} units sold",
                        'icon': 'trophy'
                    })

                if revenue_by_date:
                    best_day = max(revenue_by_date, key=lambda x: x['revenue'])
                    insights.append({
                        'type': 'top_day',
                        'title': 'Highest Revenue Day',
                        'value': best_day['date'],
                        'detail': f"${best_day['revenue']:,.2f} revenue",
                        'icon': 'calendar'
                    })

                if low_stock_alerts:
                    high_risk = [a for a in low_stock_alerts if a['risk_level'] == 'high']
                    if high_risk:
                        insights.append({
                            'type': 'inventory_risk',
                            'title': 'Inventory Risk Alert',
                            'value': f"{len(high_risk)} product(s)",
                            'detail': f"High velocity: {', '.join(a['product'] for a in high_risk[:3])}",
                            'icon': 'alert',
                            'severity': 'warning'
                        })

                if revenue_by_product:
                    insights.append({
                        'type': 'top_revenue',
                        'title': 'Top Revenue Product',
                        'value': revenue_by_product[0]['product'],
                        'detail': f"${revenue_by_product[0]['revenue']:,.2f}",
                        'icon': 'dollar'
                    })

                upload_time = datetime.fromtimestamp(
                    upload['upload_time_utc'], tz=timezone.utc
                ).isoformat()

                return {
                    'success': True,
                    'upload_id': upload_id,
                    'filename': upload['filename'],
                    'upload_time': upload_time,
                    'summary': {
                        'total_records': upload['total_records'],
                        'total_revenue': upload['total_revenue'],
                        'total_units': upload['total_units'],
                        'active_products': len(revenue_by_product)
                    },
                    'revenue_by_product': revenue_by_product,
                    'units_by_product': units_by_product,
                    'sales_by_date': sales_by_date,
                    'revenue_by_date': revenue_by_date,
                    'low_stock_alerts': low_stock_alerts,
                    'insights': insights
                }

        except Exception as e:
            logger.error(f"Error fetching analytics: {e}", exc_info=True)
            return None

    def clear_all_data(self) -> dict:
        """
        Delete all analytics data from the database.

        Returns:
            Dict with success flag and counts of deleted rows.
        """
        try:
            with sqlite3.connect(str(self.db_path)) as conn:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM analytics_summaries")
                summaries_deleted = cursor.rowcount
                cursor.execute("DELETE FROM processed_sales")
                sales_deleted = cursor.rowcount
                cursor.execute("DELETE FROM sales_uploads")
                uploads_deleted = cursor.rowcount
                conn.commit()

            logger.info(
                f"Analytics data cleared: {uploads_deleted} uploads, "
                f"{sales_deleted} sales records, {summaries_deleted} summaries deleted"
            )
            return {
                'success': True,
                'deleted': {
                    'uploads': uploads_deleted,
                    'sales_records': sales_deleted,
                    'summaries': summaries_deleted
                }
            }
        except Exception as e:
            logger.error(f"Error clearing analytics data: {e}", exc_info=True)
            return {'success': False, 'error': str(e)}

    def get_upload_history(self, limit: int = 10) -> List[dict]:
        """
        Get recent upload history.

        Args:
            limit: Maximum number of records

        Returns:
            List of upload records
        """
        try:
            with sqlite3.connect(str(self.db_path)) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()

                cursor.execute("""
                    SELECT id, filename, upload_time_utc, total_records,
                           total_revenue, total_units, status
                    FROM sales_uploads
                    ORDER BY upload_time_utc DESC
                    LIMIT ?
                """, (limit,))

                uploads = []
                for row in cursor.fetchall():
                    uploads.append({
                        'upload_id': row['id'],
                        'filename': row['filename'],
                        'upload_time': datetime.fromtimestamp(
                            row['upload_time_utc'], tz=timezone.utc
                        ).isoformat(),
                        'total_records': row['total_records'],
                        'total_revenue': row['total_revenue'],
                        'total_units': row['total_units'],
                        'status': row['status']
                    })

                return uploads

        except Exception as e:
            logger.error(f"Error fetching upload history: {e}")
            return []
